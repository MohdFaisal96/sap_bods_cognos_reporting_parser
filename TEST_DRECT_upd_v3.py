import flask
from flask import Flask, jsonify, request
from flask import send_from_directory, send_file
from flask_cors import CORS, cross_origin
from joblib import Parallel, delayed
import sys
import shutil
from datetime import datetime
import pandas as pd
import numpy as py
import time
from collections import OrderedDict
import requests
import json
import re
import ast
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from concurrent.futures import ProcessPoolExecutor
import os
import multiprocessing
import psutil
import glob2
import zipfile
from pathlib import Path
import xlrd
import csv
import warnings

warnings.filterwarnings("ignore")
app = Flask(__name__)
CORS(app)
RootPath=os.getcwd()
print(RootPath)

def drect(username, password, hostURL, arg, auth):
    # username = request.args.get('username')
    # password = request.args.get('password')

    # username = info['username']
    print("username", username)
    # password = info['password']
    print("password", password)
    # app_link = info['app_link']
    print("hostURL", hostURL)
    RepID = arg
    print("RepID", RepID)

    print("auth", auth)

    login_link = 'http://www.sap.com/rws/bip'
    request_link = hostURL + '/biprws/logon/long'
    int_link = hostURL + '/biprws/raylight/v1/documents/'
    doc_link = hostURL + '/biprws/v1/documents/'
    doc_link_2 = hostURL + '/biprws/raylight/v1/documents/' + arg

    ##Login to BO environment

    payload = '<attrs xmlns="' + login_link + '"><attr name="userName" type="string">' + username + '</attr><attr name="password" type="string">' + password + '</attr><attr name="auth" type="string" possibilities="secEnterprise,secLDAP,secWinAD,secSAPR3">' + auth + '</attr></attrs>'

    headers = {
        'content-type': 'application/xml',
        'accept': 'application/xml'
    }

    response = requests.request("POST", request_link, headers=headers, data=payload)

    json_out = response.headers

    json_out['X-SAP-LogonToken']

    headers = {
        'content-type': 'application/x-www-form-urlencoded',
        'accept': 'application/json',
        'X-SAP-LogonToken': json_out['X-SAP-LogonToken']
    }

    ##Convert list to string
    def listToString(s, e=","):
        str1 = e
        return (str1.join(s))

        ##BO initial request link

    # int_link=read_json('int_link')

    ##List Of Document
    response = requests.request("GET", doc_link, headers=headers, data=payload)
    # print('response 11:', response)

    # print("doc_link_2",doc_link_2)
    # print("headers",headers)
    # print("payload",payload)

    response_2 = requests.request("GET", doc_link_2, headers=headers, data=payload)
    # print('response doc link_2:', response_2)

    def extract_json_docid(response, module, key1, key2, value):
        response = response
        result = []
        response = OrderedDict(json.loads(response))
        if (response[module][key1] == int(key2)):
            result.append(response[module][value])
        return result

    ##Get the BO report id on passing CUID

    doc_list = [RepID]

    # cuid = extract_json_docid(response.text, 'entries', 'id', RepID, 'cuid')

    print('doc_list:', doc_list)
    rank_doc_list = doc_list


    ##Get the BO report name
    

    rep_name = extract_json_docid(response_2.text, 'document', 'id', RepID, 'name')


    ##Create a Pandas Excel writer using XlsxWriter as the engine.
    try:
        if not os.path.exists('output/' + rep_name[0] + '/'):
            os.makedirs('output/' + rep_name[0] + '/')
        writer = pd.ExcelWriter('output/' + rep_name[0] + '/' + rep_name[0] + '_BO_DFT.xlsx', engine='xlsxwriter')
        output_excel = 'output/' + rep_name[0] + '/' + rep_name[0] + '_BO_DFT.xlsx'

        ##Response genertated from BO
        def get_response(in_list, link_str, flag, link1):
            response1 = []
            link_final = []
            if flag == 0:
                for i in in_list:
                    link = link1 + i + "/" + link_str
                    link1 = link
                    response = requests.request("GET", link1, headers=headers, data=payload)
                    response = response.text
                    return response, link
            else:

                for i in in_list:

                    if link_str == '':
                        link = link1 + "/" + i

                    else:
                        link = link1 + "/" + i + "/" + link_str

                    response = requests.request("GET", link, headers=headers, data=payload)
                    response = response.text

                    response1.append(response)
                    link_final.append(link)

                return response1, link_final

        print("Pinging Server For Response...")

        ##Get position from replacements of strings

        def give_position(input):
            start = [m.start() for m in re.finditer("\[", input)]
            end = [m.start() for m in re.finditer("\]", input)]
            requiredLocs = []
            requiredLoce = []
            flag = 0
            for i in range(len(end)):
                e = end[i]
                if (flag == 1):
                    flag = 0;
                    continue;
                elif (e == len(input) - 1):
                    requiredLocs.append(start[i])
                    requiredLoce.append(e)
                    flag = -1
                elif (input[e + 1] == '.'):
                    flag = 1
                    continue
                else:
                    requiredLocs.append(start[i])
                    requiredLoce.append(e)
            if len(requiredLoce) == 0:
                return None, None, None
            if len(requiredLoce) == 1:
                flag = -1
            else:
                flag = 0
            return (requiredLocs[0], requiredLoce[0], flag)

    ############################################################################# dataProvider Starts ############################################################################

    

        ##Get the dataProvider Name for each FormulaLanguageID
        def get_dataProviderName(input, df_dp_list):

            if '[' in input:
                if give_position(input)[0] == None:
                    return input

                else:
                    requiredLocs = give_position(input)[0]
                    requiredLoce = give_position(input)[1]
                    end_indicator = give_position(input)[2]

                    while (True):
                        if end_indicator == None:
                            return input
                            break

                        replace_str = input[requiredLocs:requiredLoce + 1]

                        for item in df_dp_list:
                            replace_item = item
                            if replace_str == item.split('.')[1]:
                                input = input.replace(replace_str, replace_item)

                                break
                            elif replace_str == item:
                                input = input
                                break
                            else:
                                input = input

                        if end_indicator != -1:
                            requiredLocs = give_position(input)[0]
                            requiredLoce = give_position(input)[1]
                            end_indicator = give_position(input)[2]
                        else:
                            return input
                            break

            else:
                return input

        ##Dataproviders details
        # try:
        link = get_response(doc_list, 'dataproviders', 0, int_link)[1]
        response = OrderedDict(json.loads(get_response(doc_list, 'dataproviders', 0, int_link)[0]))
        suproya = link
        df_dp = pd.DataFrame.from_dict(response['dataproviders']['dataprovider'])
        dp_list = []
        dp_list = df_dp['id'].tolist()
        link1 = link
        flag = 1
        counter = 0
        print(link)
        response = []
        response = get_response(dp_list, '', 1, link1)[0]
        chandan = response
        print('response_1:', response)
        for i in response:
            # print('i1:',i)
            response = OrderedDict(json.loads(i))
            # print('response_2:',response)
            df_dp_temp = pd.DataFrame.from_dict(response['dataprovider'])
            abc = df_dp_temp
            df_dp_temp_sheet = df_dp_temp
            # print('df_dp_temp_sheet cuid:',df_dp_temp_sheet['dataSourceCuid'])
            # print('df_dp_temp_sheet:',df_dp_temp_sheet,' : cols : ',df_dp_temp_sheet.columns)
            # print('df_dp_temp_sheet name col:',df_dp_temp_sheet['name'])

            counter += 1
            df_dp_temp = df_dp_temp[df_dp_temp.index == 'expression']
            df_dp_temp['Sheet Name'] = None
            df_dp_temp['File Name'] = None
            df_dp_temp['Custom Query Indicator'] = ''
            df_dp_temp['Universe Name'] = ''
            # print('df_dp_temp:',df_dp_temp)
            # print('df_dp_temp["dataSourceType"]:',df_dp_temp['dataSourceType'])
            ##Cavet to handle datasource type 'excel'
            if df_dp_temp['dataSourceType'].iloc[0] == 'excel':
                df_dp_temp.rename(columns={'properties': 'query'}, inplace=True)
                df_dp_temp = df_dp_temp.drop(columns=['dataSourcePrefix'])
                ##Addded to get sheet name
                df_dp_temp_sheet = df_dp_temp_sheet[df_dp_temp_sheet.index == 'property']
                # print('df_dp_temp_sheet 2:',df_dp_temp_sheet)
                cuid_excel = (df_dp_temp_sheet.loc['property']['dataSourceCuid'])
                # print('cuid_excel::',cuid_excel)
                # print('name::',df_dp_temp_sheet.loc['property']['name'])
                # print('excel:',df_dp_temp['dataSourceType'].iloc[0],'  ',df_dp_temp['dataSourceType'])
                # print('file name :',extract_json_docid(response,'dataprovider','dataSourceCuid',cuid_excel,'name'))
                # df_dp_temp['File Name']=extract_json_docid(response,'dataprovider','dataSourceCuid',cuid_excel,'name')
                # df_dp_temp['File Name'] = extract_json_docid(response_text, 'entries', 'cuid', cuid_excel, 'name')
                df_dp_temp['File Name'] = df_dp_temp_sheet.loc['property']['name']
                df_dp_temp['Sheet Name'] = (df_dp_temp_sheet.loc['property']['properties'])[0]['$']
            else:
                # print('not excel:',df_dp_temp['dataSourceType'].iloc[0],'  ',df_dp_temp['dataSourceType'])
                pass
                # df_dp_temp['Sheet Name'].iloc[counter]=None

            ##Apending to final dataframe all dataprovider details

            if flag == 1:
                df_dp_final = df_dp_temp
                flag = 0
            else:
                df_dp_final = df_dp_final.append(df_dp_temp)

        flag = 1
        for k, v in df_dp_final['dictionary'].items():
            df = pd.DataFrame.from_dict(v)
            if 'aggregationFunction' in list(df.columns):
                df = df[['id', 'formulaLanguageId', 'name', 'dataSourceObjectId', 'aggregationFunction']]
            else:
                df = df[['id', 'formulaLanguageId', 'dataSourceObjectId', 'name']]
                df['aggregationFunction'] = None
            if flag == 1:
                df_dp_detail = df
                flag = 0
            else:
                df_dp_detail = df_dp_detail.append(df, ignore_index=True)

        df_dp_final = df_dp_final[
            ['id', 'name', 'dataSourceType', 'query', 'Sheet Name', 'File Name', 'Custom Query Indicator',
             'Universe Name']]

        ##Rename to avoid namespace clash while joining and better readeability
        df_dp_final['name'] = '[' + df_dp_final['name'] + ']'
        df_dp_final.rename(columns={'name': 'DataProvider Name'}, inplace=True)
        df_dp_detail = df_dp_detail[['id', 'formulaLanguageId', 'name', 'dataSourceObjectId', 'aggregationFunction']]
        df_dp_detail['pid'] = df_dp_detail['id'].str.split('.').str[0]

        ##Rename to avoid namespace clash while joining and better readeability
        df_dp_detail.rename(columns={'id': 'sid'}, inplace=True)
        df_dp_final = pd.merge(df_dp_final, df_dp_detail, how='right', left_on=['id'], right_on=['pid'])
        df_dp_final = df_dp_final[
            ['pid', 'dataSourceObjectId', 'dataSourceType', 'File Name', 'DataProvider Name', 'Sheet Name', 'query',
             'sid',
             'formulaLanguageId', 'name', 'aggregationFunction', 'Custom Query Indicator',
             'Universe Name']]  # 1st tab is written
        df_dp_final['formulaLanguageId'] = py.where(
            df_dp_final['formulaLanguageId'].str.contains(re.compile("\.")) == True,
            df_dp_final['formulaLanguageId'],
            df_dp_final['DataProvider Name'] + '.' + df_dp_final[
                'formulaLanguageId'])

        ##Rename to avoid namespace clash while joining and better readeability
        df_dp_final.rename(columns={'sid': 'id'}, inplace=True)
        tempdp_list = df_dp_final['DataProvider Name'].tolist()
        for i in range(len(tempdp_list)):
            tempdp_list[i] = tempdp_list[i].replace('[', '').replace(']', '')
        df_dp_final['DataProvider Name'] = tempdp_list
        custom_query_res = requests.request("GET", link, headers=headers, data=payload)
        custom_query_res = OrderedDict(json.loads(custom_query_res.text))
        df_custom_temp = pd.DataFrame.from_dict(custom_query_res['dataproviders']['dataprovider'])
        custom_query_dict = {}
        for index in df_custom_temp.index:
            if df_custom_temp['dataSourceType'][index] == 'unv' or df_custom_temp['dataSourceType'][index] == 'unx':
                custom_query_link = link + '/' + df_custom_temp['id'][index] + '/queryplan'
                custom_query_res = requests.request("GET", custom_query_link, headers=headers, data=payload)
                custom_query_res = OrderedDict(json.loads(custom_query_res.text))
                custom_query_dict[df_custom_temp['id'][index]] = custom_query_res.get('queryplan').get('@custom')

        for key in custom_query_dict:
            for index in df_dp_final.index:
                if df_dp_final['pid'][index] == key:
                    df_dp_final['Custom Query Indicator'][index] = custom_query_dict[key]

        # Get universe name
        universe_res = requests.request("GET", link, headers=headers, data=payload)
        universe_res = OrderedDict(json.loads(universe_res.text))
        df_universe_temp = pd.DataFrame.from_dict(universe_res['dataproviders']['dataprovider'])
        universe_dict = {}
        for index in df_universe_temp.index:
            if df_universe_temp['dataSourceType'][index] == 'unv' or df_universe_temp['dataSourceType'][index] == 'unx':
                universe_link = "http://cvyhj1a18:6405/biprws/raylight/v1/universes/" + \
                                df_universe_temp['dataSourceId'][index]
                universe_res = requests.request("GET", universe_link, headers=headers, data=payload)
                universe_res = OrderedDict(json.loads(universe_res.text))
                universe_dict[df_universe_temp['id'][index]] = universe_res.get('universe').get('name')

        for key in universe_dict:
            for index in df_dp_final.index:
                if df_dp_final['pid'][index] == key:
                    df_dp_final['Universe Name'][index] = universe_dict[key]

        df_dp_final_bkp1 = df_dp_final
        df_dp_final.to_excel(writer, sheet_name='DataProvider', index=False)

        ##Get the data Providers
        # df_dp_list=df_dp_final['formulaLanguageId'].tolist()

    ############################################################################# dataProvider Ends ############################################################################


    ############################################################################# Calculations Starts ############################################################################


        ##Get list of variable
        link = get_response(doc_list, 'variables', 0, int_link)[1]
        response = OrderedDict(json.loads(get_response(doc_list, 'variables', 0, int_link)[0]))
        var_list = []

        if len(response['variables']['variable']) == 0:  ##No variable associated with the report
            df_var = pd.DataFrame(columns=['id', 'Calculated Field Name', 'DataProvider', 'Formula', 'Qualification'])
            df_var.to_excel(writer, sheet_name='Calculations', index=False)
        else:
            for doc in response['variables']['variable']:  ##Variable associated with the report
                var_list.append(doc['id'])
            ###Get details of each variable
            flag = 1
            for i in var_list:
                link1 = link + "/" + i
                response = requests.request("GET", link1, headers=headers, data=payload)
                # df=pd.read_json(response.text,orient='records').transpose()
                if flag == 1:
                    df = pd.read_json(response.text, orient='records').transpose()
                    df1 = df
                    flag = 0
                else:
                    df = pd.read_json(response.text, orient='records').transpose()
                    df1 = df1.append(df)
                    #df1 = pd.concat(df1,df)

            df_var = pd.DataFrame(columns=['id', 'Calculated Field Name', 'DataProvider', 'Formula', 'Qualification'])
            df_var['id'] = df1['id']
            df_var['Calculated Field Name'] = df1['name']
            df_var['Formula'] = df1['definition']
            df_var['Qualification'] = df1['@qualification']
            df_var = df_var.replace(regex=r'^=', value='')
            df_cal_query = df_dp_final_bkp1
            formula_list = df_var['Formula'].tolist()
            print(formula_list)
            print(len(formula_list))
            queryname = []
            randomquery = df_cal_query['DataProvider Name'].iloc[0]
            for formula in formula_list:
                if '].[' in formula:
                    query = formula[::-1].split('[.]')[1].split('[')[0]
                    query = '[' + query[::-1] + ']'
                    queryname.append(query.replace('[', '').replace(']', ''))
                    randomquery = query
                elif '[' in formula:
                    query = formula.split('[')[1].split(']')[0]
                    dftemp = df_cal_query.loc[df_cal_query['name'] == query]
                    print('dftemp:', dftemp['DataProvider Name'])
                    query = dftemp['DataProvider Name'].values[0]
                    queryname.append(query.replace('[', '').replace(']', ''))
                    randomquery = query
                else:
                    query = randomquery
                    queryname.append(query.replace('[', '').replace(']', ''))
            df_var['DataProvider'] = queryname
            df_var.to_excel(writer, sheet_name='Calculations', index=False)
        df_var['IsMerge'] = None  ##Created for report element details

         ############################################################################# Calculations Ends ############################################################################

        ##Get list of links
        link = get_response(doc_list, 'links', 0, int_link)[1]
        response = OrderedDict(json.loads(get_response(doc_list, 'links', 0, int_link)[0]))
        link_list = []

     ############################################################################# Merge Start ############################################################################

        ##Get MergeDimentions
        def get_MergeDimension(input):
            a = input
            b = ''
            for i in range(len(a['linkedExpression'])):
                b = b + str(a['linkedExpression'][i].get('@id')) + ','
            return (listToString(b.split(',')[0:-1]))

        ##Get Merge Dimention name
        def get_formulaLanguageId(input):
            a = input
            out = []
            b = a.split(',')
            df = df_dp_final[['id', 'formulaLanguageId']]
            for i in b:
                out.append(listToString(df['formulaLanguageId'][df['id'] == i].values.tolist()))
            return listToString(out)

        if len(response['links']['link']) == 0:  ##No MergeDimentions associated with the report
            df_link = pd.DataFrame(
                columns=['id', 'Calculated Field Name', '@dataType', '@qualification', 'dataSourceObjectId',
                         'formulaLanguageId', 'MergeDimensionID', 'MergeDimension'])
            df_link.to_excel(writer, sheet_name='Merge_Dim', index=False)
        else:
            for doc in response['links']['link']:  ##Merge Dimensions associated with the report
                link_list.append(doc['id'])

            ###Get details of each links
            flag = 1
            for i in link_list:
                link1 = link + "/" + i
                response = requests.request("GET", link1, headers=headers, data=payload)

                if flag == 1:
                    df = pd.read_json(response.text, orient='records').transpose()
                    df1 = df
                    flag = 0
                else:
                    df = pd.read_json(response.text, orient='records').transpose()
                    df1 = df1.append(df)
            df_link = df1.replace(regex=r'^=', value='')
            df_link = df_link[['id', 'name', '@dataType', '@qualification', 'dataSourceObjectId', 'formulaLanguageId',
                               'linkedExpressions']]
            df_link.rename(columns={'name': 'Calculated Field Name', 'linkedExpressions': 'MergeDimensionID'},
                           inplace=True)
            df_link['MergeDimension'] = None
            i = 0
            for i in range(df_link.shape[0]):
                if (df_link['MergeDimensionID'].iloc[i]) != None:
                    merge_id = get_MergeDimension(df_link['MergeDimensionID'].iloc[i])
                    df_link['MergeDimensionID'].iloc[i] = merge_id
                    df_link['MergeDimension'].iloc[i] = get_formulaLanguageId(merge_id)
            df_link.to_excel(writer, sheet_name='Merge_Dim', index=False)
        df_link['IsMerge'] = 'MergeDimension'  ##Created for report element detail

     ############################################################################# Merge End ############################################################################

        ##Get report id list
        link = get_response(doc_list, 'reports', 0, int_link)[1]
        response = json.loads(get_response(doc_list, 'reports', 0, int_link)[0])

        rep_list = {}
        for doc in response['reports']['report']:
            rep_list[doc['id']] = doc['name']

        ############################################################################# Driller Start ############################################################################

        ##Get Data Driller Details
        flag1 = 1
        for i, j in rep_list.items():
            link1 = link + "/" + str(i)
            response = requests.request("GET", link1, headers=headers, data=payload)
            response = json.loads(response.text)
            response = OrderedDict(response)
            df = pd.DataFrame.from_dict(response)
            df = df.transpose()
            df = df[['@hasDatafilter', '@hasDriller', 'id', 'name']]
            if flag1 == 1:
                df_driller = df
                flag1 = 0
            else:
                df_driller = df_driller.append(df, ignore_index=True)

        df_rep_filter = df_driller[df_driller['@hasDatafilter'] == 'true']
        df = df_driller[df_driller['@hasDriller'] == 'true']

        i = 0
        if (df.shape[0]) == 0:
            df_driller = pd.DataFrame(
                columns=['Report ID', 'Report Name', 'DataSourceID', 'name', 'Qualification', 'value'])
            df_driller.to_excel(writer, sheet_name='DrillerDetails', index=False)
        else:
            flag = 1
            while i in range(df.shape[0]):
                link2 = link + "/" + str(df['id'].iloc[i]) + "/driller/filters/"
                response = requests.request("GET", link2, headers=headers, data=payload)
                response = json.loads(response.text)
                response = OrderedDict(response)
                rep_id = str(df['id'].iloc[i])
                rep_name = str(df['name'].iloc[i])
                ##Driller enabled but no filters assigned
                if len(response['filters']['filter']) == 0:
                    df1 = pd.DataFrame(index=range(0, 1),
                                       columns=['Report ID', 'Report Name', 'id', 'name', '@qualification', 'value'])
                    df1['Report ID'] = rep_id
                    df1['Report Name'] = rep_name
                else:
                    df1 = pd.DataFrame.from_dict(response['filters']['filter'])
                    df1['Report ID'] = rep_id
                    df1['Report Name'] = rep_name
                    if 'value' in df1.columns:
                        df1 = df1[['Report ID', 'Report Name', 'id', 'name', '@qualification', 'value']]
                    else:
                        df1['value'] = None
                        df1 = df1[['Report ID', 'Report Name', 'id', 'name', '@qualification', 'value']]
                if flag == 1:
                    df_driller = df1
                    flag = 0
                else:
                    df_driller = df_driller.append(df1, ignore_index=True)
                i = i + 1
            df_driller.rename(columns={'id': 'DataSourceID', '@qualification': 'Qualification'}, inplace=True)
            df_driller.to_excel(writer, sheet_name='DrillerDetails', index=False)

        ############################################################################# Driller END ############################################################################
        flag = 1
        flag3 = 1

        df_rep_final = pd.DataFrame(columns=['rep_id', 'id', 'name', 'type'])
        for i, j in rep_list.items():

            link1 = link + "/" + str(i) + "/elements"

            response = requests.request("GET", link1, headers=headers, data=payload)
            response = json.loads(response.text)
            response = OrderedDict(response)
            df = pd.DataFrame.from_dict(response)
            len_elm = len(df['elements']['element'])
            n = 0
            dict1 = {}
            df_rep = pd.DataFrame(index=range(0, len_elm), columns=['rep_id', 'id', 'name', 'type'])

            while n < len_elm:
                dict1 = df['elements']['element'][n]
                for k, v in dict1.items():

                    if k == 'id':
                        df_rep['id'].iloc[n] = v
                    if k == '@type':
                        df_rep['type'].iloc[n] = v
                    if k == 'name':
                        df_rep['name'].iloc[n] = v

                n = n + 1
            df_rep['rep_id'] = i
            df_rep['rep_name'] = j
            list_element = []
            temp = ['Visualization', 'VTable', 'XTable', 'HTable']
            # temp=['XTable']
            list_element = df_rep[df_rep['type'].isin(temp)]['id'].to_list()

            flag1 = 1

            for p in list_element:
                link2 = link1 + "/" + str(p)

                response = requests.request("GET", link2, headers=headers, data=payload)
                response = json.loads(response.text)
                response = OrderedDict(response)
                df2 = pd.DataFrame.from_dict(response).transpose()
                df2['rep_id'] = i
                df2['rep_name'] = j
                if flag1 == 1 & flag == 1:
                    df_br = df2
                    flag1 = 0
                    flag = 0
                else:
                    df_br = df_br.append(df2, ignore_index=True)

            df_rep_final = df_rep_final.append(df_rep, ignore_index=True)

        df_br.sort_values(by=['rep_id'], inplace=True)

        i = 0
        flag2 = 1
        ##Works only for table
        while i < df_br.shape[0]:

            if df_br['@type'].iloc[i] in ['VTable', 'XTable', 'HTable']:
                df3 = df_br['content'].iloc[i]
                counter = len(df3['axes']['axis'])
                k = 0
                while k < counter:
                    df4 = pd.DataFrame.from_dict(df3['axes']['axis'][k]['expressions']['formula'])
                    def_element = pd.DataFrame(index=range(0, df4.shape[0]),
                                               columns=['Requirement ID', 'Report ID', 'Report Name', 'Chart Title',
                                                        'Chart Type', 'Query Name', 'hyper link', 'Field Name',
                                                        'Qualification', 'Aggregation', 'Filter Applicable'])
                    def_element['Chart Title'] = df_br['name'].iloc[i]
                    def_element['EID'] = df_br['id'].iloc[i]
                    def_element['Report ID'] = df_br['rep_id'].iloc[i]
                    def_element['Report Name'] = df_br['rep_name'].iloc[i]
                    def_element['Chart Type'] = df_br['@type'].iloc[i]
                    def_element['Requirement ID'] = 'BR00' + str(i + 1)
                    def_element['Filter Applicable'] = df_br['@hasDatafilter'].iloc[i]

                    def_element['Qualification'] = df4['@qualification']

                    def_element = def_element.replace(regex='[^\w]', value=' ')
                    def_element = def_element.replace(regex='^ ', value='')
                    def_element['Aggregation'] = df4['@dataObjectId']
                    def_element['formulaLanguageId'] = df4['$']
                    def_element['hyper link'] = df4[df4['$'].str.contains('href')]['$']
                    def_element = def_element.replace(regex=r'^=', value='')
                    k = k + 1
                    if flag2 == 1 & flag3 == 1:
                        def_element_final = def_element
                        flag2 = 0
                        flag3 = 0
                    else:
                        def_element_final = def_element_final.append(def_element, ignore_index=True)
                i = i + 1
            else:

                df3 = df_br['content'].iloc[i]
                df4 = pd.DataFrame.from_dict(df3['chart']['axes']['axis'])
                df4 = df4[df4['expressions'].notna()]
                df4 = pd.DataFrame.from_dict(df4['expressions'])
                j = 0
                dict_ele = {}
                f1 = 0
                while j < df4.shape[0]:
                    dict_ele = df4.iloc[j][0]
                    df5 = pd.DataFrame.from_dict(dict_ele['formula'])
                    dict_ele = {}
                    if f1 == 0:
                        df6 = df5
                        f1 = 1
                    else:
                        df6 = df6.append(df5, ignore_index=True)

                    j = j + 1
                def_element = pd.DataFrame(index=range(0, df6.shape[0]),
                                           columns=['Requirement ID', 'Report ID', 'EID', 'Report Name', 'Chart Title',
                                                    'Chart Type', 'Query Name', 'hyper link', 'Field Name',
                                                    'Qualification',
                                                    'Aggregation', 'Filter Applicable'])
                def_element['Chart Title'] = df_br['name'].iloc[i]
                def_element['EID'] = df_br['id'].iloc[i]
                def_element['Requirement ID'] = 'BR00' + str(i + 1)
                def_element['Report ID'] = df_br['rep_id'].iloc[i]
                def_element['Report Name'] = df_br['rep_name'].iloc[i]
                def_element['Chart Type'] = df3['chart']['@type']
                def_element['Filter Applicable'] = df_br['@hasDatafilter'].iloc[i]

                def_element['Qualification'] = df6['@qualification']

                def_element = def_element.replace(regex='[^\w]', value=' ')
                def_element = def_element.replace(regex='^ ', value='')
                def_element['Aggregation'] = df6['@dataObjectId']
                def_element['formulaLanguageId'] = df6['$']
                def_element = def_element.replace(regex=r'^=', value='')
                if flag2 == 1 & flag3 == 1:
                    def_element_final = def_element
                    flag2 = 0
                    flag3 = 0
                else:
                    def_element_final = def_element_final.append(def_element, ignore_index=True)
                i = i + 1

        df_dp_final = df_dp_final[['id', 'aggregationFunction', 'name', 'DataProvider Name']]

        def_element_final = pd.merge(def_element_final, df_dp_final, how='left', left_on=['Aggregation'],
                                     right_on=['id'])
        def_element_final['Field Name'] = def_element_final['name']
        df_exp = pd.concat(
            [df_var[['id', 'Calculated Field Name', 'IsMerge']], df_link[['id', 'Calculated Field Name', 'IsMerge']]],
            ignore_index=True)

        def_element_final = pd.merge(def_element_final, df_exp[['id', 'Calculated Field Name', 'IsMerge']], how='left',
                                     left_on=['Aggregation'], right_on=['id'])
        def_element_final['Field Name'] = py.where(def_element_final['name'].astype(str) == 'nan',
                                                   def_element_final['Calculated Field Name'],
                                                   def_element_final['name'])
        def_element_final['Query Name'] = def_element_final['formulaLanguageId']
        def_element_final['Aggregation'] = def_element_final['aggregationFunction']

        def_element_final = def_element_final[
            ['Requirement ID', 'Report ID', 'EID', 'Report Name', 'Chart Title', 'Chart Type', 'DataProvider Name',
             'Query Name', 'hyper link', 'Field Name', 'IsMerge', 'Qualification', 'Aggregation', 'Filter Applicable']]
        def_element_final.sort_values(by=['Report ID', 'EID'], inplace=True)

        ##Get the dataprovider name where applicable
        # for i in range(def_element_final.shape[0]):
        #    if def_element_final['IsMerge'].iloc[i]=='MergeDimension':
        #        pass
        #    else:
        #        def_element_final['Query Name'].iloc[i]=get_dataProviderName(def_element_final['Query Name'].iloc[i],df_dp_list)

        # [print(get_dataProviderName(def_element_final['Query Name'].iloc[i],df_dp_list)) for i in range(def_element_final.shape[0])]

######################################################################## Filter Details ################################################################################

        ##Get datafilter for each object

        df_br_filter = df_br[['rep_id', 'id', '@hasDatafilter']][df_br['@hasDatafilter'] == 'true']

        if (df_br_filter.shape[0]) == 0:
            def_element_final['filter'] = None
            df_filter = pd.DataFrame(
                columns=['Report ID', 'EID', 'Report Name', 'Chart Title', 'filter', 'DataProvider Name', 'operation',
                         'value', 'Requirement ID'])
        else:
            df_filter = pd.DataFrame(columns=['e_rep_id', 'e_id', 'filter', 'operation', 'value'])
            i = 0
            while i < df_br_filter.shape[0]:
                link = get_response(doc_list, 'reports', 0, int_link)[1] + "/" + str(
                    df_br_filter['rep_id'].iloc[i]) + "/elements/" + str(df_br_filter['id'].iloc[i]) + '/datafilter'
                response = requests.request("GET", link, headers=headers, data=payload)
                response = json.loads(response.text)
                response = OrderedDict(response)
                if listToString(response['datafilter'].keys()) == 'condition':
                    df['filter'] = response['datafilter']['condition']['@key']
                    df['value'] = listToString(response['datafilter']['condition']['value'], e='~')
                    df['operation'] = response['datafilter']['condition']['@operator']
                    df['e_rep_id'] = df_br_filter['rep_id'].iloc[i]
                    df['e_id'] = df_br_filter['id'].iloc[i]
                    df_filter = df_filter.append(df, ignore_index=True)
                    ifxxx = df_filter
                if listToString(response['datafilter'].keys()) in ['and', 'or']:
                    a = listToString(response['datafilter'].keys())
                    for res in response['datafilter'][a]['condition']:
                        df['filter'] = res['@key']
                        df['value'] = listToString(res['value'], e='~')
                        df['operation'] = res['@operator']
                        df['e_rep_id'] = df_br_filter['rep_id'].iloc[i]
                        df['e_id'] = df_br_filter['id'].iloc[i]
                        df_filter = df_filter.append(df, ignore_index=True)

                i = i + 1

            # df_filter['filter']=df_filter['filter'].replace(regex='[^\w]', value=' ')
            # df_filter['filter']=df_filter['filter'].replace(regex='^ ', value='')
            df = df_filter[['e_rep_id', 'e_id', 'filter']].groupby(['e_rep_id', 'e_id'], as_index=False).agg(
                lambda x: ",".join(x))

            def_element_final = pd.merge(def_element_final, df, how='left', right_on=['e_rep_id', 'e_id'],
                                         left_on=['Report ID', 'EID'])
            df_filter = pd.merge(
                def_element_final[['Report ID', 'EID', 'Report Name', 'Chart Title', 'DataProvider Name',
                                   'Requirement ID']], df_filter, how='right',
                right_on=['e_rep_id', 'e_id'], left_on=['Report ID', 'EID'])
            xxxxx = df_filter
            def_element_final = def_element_final[
                ['Requirement ID', 'EID', 'Report ID', 'Report Name', 'Chart Title', 'Chart Type', 'DataProvider Name',
                 'Query Name', 'hyper link', 'Field Name', 'IsMerge', 'Qualification', 'Aggregation', 'filter']]
            df_filter = df_filter[
                ['Report ID', 'EID', 'Report Name', 'Chart Title', 'filter', 'DataProvider Name', 'operation', 'value',
                 'Requirement ID']].drop_duplicates()
            # df_filter=df_filter[['Report ID','EID','Report Name','Chart Title','filter','operation','value','Requirement ID']].drop_duplicates()
            xxxxx1 = df_filter
            # def_element_final.to_excel(writer,sheet_name='Worksheets',index=False)

        ##Checking for Report level filter
        if df_rep_filter.shape[0] == 0:
            if df_filter.shape[0] == 0:
                df_filter = pd.DataFrame(
                    columns=['Report ID', 'EID', 'Report Name', 'Chart Title', 'filter', 'DataProvider Name',
                             'operation',
                             'value', 'Requirement ID'])
                def_element_final['DataProvider Name'] = def_element_final['DataProvider Name'].astype(str).str.replace(
                    r"[\[\]']", '')
                # def_element_final.to_excel(writer,sheet_name='Worksheets',index=False)
                df_filter = df_filter.dropna()
                df_filter['DataProvider Name'] = df_filter['DataProvider Name'].astype(str).str.replace(r"[\[\]']", '')
                df_filter.to_excel(writer, sheet_name='FilterDetails', index=False)

            else:
                def_element_final['DataProvider Name'] = def_element_final['DataProvider Name'].astype(str).str.replace(
                    r"[\[\]']", '')
                # def_element_final.to_excel(writer,sheet_name='Worksheets',index=False)
                df_filter = df_filter.dropna()
                df_filter['DataProvider Name'] = df_filter['DataProvider Name'].astype(str).str.replace(r"[\[\]']", '')
                df_filter.to_excel(writer, sheet_name='FilterDetails', index=False)



        else:
            ##Inserting Report Level Filter Details
            link1 = get_response(doc_list, 'reports', 0, int_link)[1]
            if df_filter.shape[0] == 0:
                df_filter = pd.DataFrame(
                    columns=['Report ID', 'EID', 'Report Name', 'Chart Title', 'filter', 'DataProvider Name',
                             'operation',
                             'value', 'Requirement ID'])

            i = 0
            print('df_rep_filter shape:', df_rep_filter.shape[0])
            while i < df_rep_filter.shape[0]:
                response = requests.request("GET", link1 + '/' + str(df_rep_filter['id'].iloc[i]) + "/datafilter",
                                            headers=headers, data=payload)
                response = json.loads(response.text)
                response = OrderedDict(response)
                print('response new1:', response)
                if listToString(response['datafilter'].keys()) == 'condition':
                    print('1st if')
                    df['filter'] = response['datafilter']['condition']['@key']
                    df['value'] = listToString(response['datafilter']['condition']['value'], e='~')
                    df['operation'] = response['datafilter']['condition']['@operator']
                    df['Report ID'] = df_rep_filter['id'].iloc[i]
                    df['Report Name'] = df_rep_filter['name'].iloc[i]
                    df_filter = df_filter.append(df, ignore_index=True)
                if listToString(response['datafilter'].keys()) in ['and', 'or']:
                    a = listToString(response['datafilter'].keys())
                    print('2nd if')
                    if 'condition' in response['datafilter'][a].keys():
                        for res in response['datafilter'][a]['condition']:
                            df['filter'] = res['@key']
                            df['value'] = listToString(res['value'], e='~')
                            df['operation'] = res['@operator']
                            df['Report ID'] = df_rep_filter['id'].iloc[i]
                            df['Report Name'] = df_rep_filter['name'].iloc[i]
                            df_filter = df_filter.append(df, ignore_index=True)
                i = i + 1

            df_filter = df_filter[
                ['Report ID', 'EID', 'Report Name', 'Chart Title', 'filter', 'DataProvider Name', 'operation', 'value',
                 'Requirement ID']].drop_duplicates()

            # df_filter=df_filter[['Report ID','EID','Report Name','Chart Title','filter','operation','value','Requirement ID']].drop_duplicates()
            ##Added to get the BuildRequirement Level Details for Report level filter sheet
            df_filter_rl = df_filter[df_filter['EID'].isnull()]
            df_filter_rl = pd.merge(
                def_element_final[['Requirement ID', 'Report ID', 'EID', 'Report Name', 'Chart Title',
                                   'DataProvider Name']],
                df_filter_rl[['Report ID', 'filter', 'operation', 'value']], how='right',
                right_on=['Report ID'], left_on=['Report ID']).drop_duplicates()
            df_filter = pd.concat([df_filter, df_filter_rl], ignore_index=True)
            df_filter = df_filter[df_filter['EID'].isnull() == False]
            df_filter = df_filter[
                ['Report ID', 'EID', 'Report Name', 'Chart Title', 'filter', 'DataProvider Name', 'operation', 'value',
                 'Requirement ID']]
            ##Added to get the report level filter in worksheet tab
            df_filter_rl.rename(columns={'filter': 'f_column'}, inplace=True)
            def_element_final = pd.merge(def_element_final, df_filter_rl[['Report ID', 'f_column']], how='left',
                                         right_on=['Report ID'], left_on=['Report ID'])
            def_element_final['filter'] = py.where(def_element_final['filter'].astype(str).isin(['nan', 'None']),
                                                   def_element_final['f_column'], def_element_final['filter'])
            # def_element_final['filter']=py.where(def_element_final['filter'].astype(str)=='None',def_element_final['f_column'],def_element_final['filter'])
            def_element_final = def_element_final[
                ['Requirement ID', 'EID', 'Report ID', 'Report Name', 'Chart Title', 'Chart Type', 'DataProvider Name',
                 'Query Name', 'hyper link', 'Field Name', 'IsMerge', 'Qualification', 'Aggregation',
                 'filter']].drop_duplicates()
            # df_filter=df_filter[df_filter['Dataprovider Name']!=None]
            def_element_final['DataProvider Name'] = def_element_final['DataProvider Name'].astype(str).str.replace(
                r"[\[\]']", '')
            df_work_sheet = def_element_final
            for idx, rows in df_work_sheet.iterrows():
                hlink = rows['hyper link']
                hlink = str(hlink)
                if "<a href" in hlink:
                    query = hlink.split('+[')[1].split(']+')[0]
                    link = hlink.split('href=')[1].split('title=')[0]
                    df_work_sheet['Field Name'].loc[idx] = query
                    df_work_sheet['hyper link'].loc[idx] = link
                else:
                    df_work_sheet['Field Name'] = def_element_final['Field Name']
            # df_work_sheet.to_excel(writer,sheet_name='Worksheets',index=False)
            df_filter = df_filter.dropna()
            df_filter['DataProvider Name'] = df_filter['DataProvider Name'].astype(str).str.replace(r"[\[\]']", '')
            print(df_filter.columns, '774')
            to_remove = ['Custom_Query']
            df_filter = df_filter[~df_filter['DataProvider Name'].isin(to_remove)]
            df_filter.to_excel(writer, sheet_name='FilterDetails', index=False)

        # Remove bracket from filter values
        w_filter_list = def_element_final['filter'].tolist()
        for i in range(len(w_filter_list)):
            if str(w_filter_list[i]) == 'nan':
                w_filter_list[i] = ''
            else:
                w_filter_list[i] = str(w_filter_list[i]).replace('[', '').replace(']', '')
        def_element_final['filter'] = w_filter_list
        f_filter_list = df_filter['filter'].tolist()
        for i in range(len(f_filter_list)):
            if str(f_filter_list[i]) == 'nan':
                f_filter_list[i] = ''
            elif '].[' in f_filter_list[i]:
                f_filter_list[i] = str(f_filter_list[i]).split('].[')[1].split(']')[0].replace('[', '').replace(']', '')
            else:
                f_filter_list[i] = str(f_filter_list[i]).replace('[', '').replace(']', '')
        df_filter['filter'] = f_filter_list
        # def_element_final.to_excel(writer,sheet_name='Worksheets',index=False)
        df_filter.to_excel(writer, sheet_name='FilterDetails', index=False)


################################################################## FilterDetails Ends ################################################################## 

        w_dataprovidername_list = def_element_final['DataProvider Name'].tolist()
        w_queryname_list = def_element_final['Query Name'].tolist()
        c_dataprovider_list = df_var['DataProvider'].tolist()
        c_calculatedfieldname_list = df_var['Calculated Field Name'].tolist()
        d_name_list = df_dp_final_bkp1['name'].tolist()
        d_dataprovidername_list = df_dp_final_bkp1['DataProvider Name'].tolist()

        for i in range(len(w_dataprovidername_list)):
            if w_dataprovidername_list[i] == 'nan':
                if '].[' in w_queryname_list[i]:
                    query = w_queryname_list[i][::-1].split('''[.]''')[1].split('''[''')[0]
                    query = '[' + query[::-1] + ']'
                    w_dataprovidername_list[i] = query.replace('[', '').replace(']', '')
                elif '[' in w_queryname_list[i]:
                    query = w_queryname_list[i].split('[')[1].split(']')[0]
                    for j in range(len(c_calculatedfieldname_list)):
                        if c_calculatedfieldname_list[j] == query:
                            w_dataprovidername_list[i] = c_dataprovider_list[j].replace('[', '').replace(']', '')
                            break
                        else:
                            for k in range(len(d_name_list)):
                                if d_name_list[k] == query:
                                    w_dataprovidername_list[i] = d_dataprovidername_list[k].replace('[', '').replace(
                                        ']',
                                        '')
                                    break
        def_element_final['DataProvider Name'] = w_dataprovidername_list
        ws_field_list = def_element_final['Field Name'].tolist()
        j = 1
        i = 0
        formula_index = []
        for index, row in def_element_final.iterrows():
            if (pd.isnull(row['Field Name'])):
                ws_field_list[i] = 'formula' + str(j)
                formula_index.append(i)
                j = j + 1
            i = i + 1
        def_element_final['Field Name'] = ws_field_list

        # def_element_final.to_excel(writer,sheet_name='Worksheets',index=False)

        ############################################################################# Ranking Start ############################################################################

        # Creating Ranking Sheet



        df_rank = pd.DataFrame(
            columns=['Requirement ID', 'Report ID', 'EID', 'Calculation', 'Top', 'Bottom', 'DataProvider Name',
                     'Based On',
                     'Ranked By'])
        rank_calc = []
        rank_based = []
        rank_rankedby = []
        rank_top = []
        rank_bottom = []
        rank_report_id = []
        rank_eid = []
        rank_req_id = []
        rank_dp = []
        print("RANK DOC LIST",rank_doc_list)
        for id in rank_doc_list:
            # print('id=',id)
            ranklink = int_link + id + '/reports'
            rank_response = requests.request("GET", ranklink, headers=headers, data=payload)
            rank_response = OrderedDict(json.loads(rank_response.text))
            # print('rank_response:',rank_response)
            df_rank_temp = pd.DataFrame.from_dict(rank_response['reports']['report'])
            # print('df_rank_temp:',df_rank_temp)
            rank_id_list = df_rank_temp['id'].tolist()
            rank_name_list = df_rank_temp['name'].tolist()
            # print('listsss',rank_id_list,rank_name_list)
            for i in range(len(rank_name_list)):
                # if rank_name_list[i]=='Rank & Alert':
                ranklink2 = ranklink + '/' + str(rank_id_list[i]) + '/elements'
                rank_response = requests.request("GET", ranklink2, headers=headers, data=payload)
                rank_response = OrderedDict(json.loads(rank_response.text))
                if 'elements' in rank_response:
                    df_rank_temp = pd.DataFrame.from_dict(rank_response['elements']['element'])
                    rank_id_list2 = df_rank_temp['id'].tolist()
                    rank_name_list2 = df_rank_temp['name'].tolist()
                    for j in range(len(rank_name_list2)):
                        # if rank_name_list2[j]=='Rank':
                        ranklink3 = ranklink2 + '/' + str(rank_id_list2[j]) + '/ranking'
                        rank_response = requests.request("GET", ranklink3, headers=headers, data=payload)
                        rank_response = OrderedDict(json.loads(rank_response.text))
                        df_rank_temp = pd.DataFrame.from_dict(rank_response.get('ranking'))
                        if df_rank_temp.empty:
                            pass

                        # df_no_rank = pd.DataFrame(columns=['Requirement ID', 'Report ID', 'EID', 'Calculation', 'Top', 'Bottom', 'DataProvider Name','Based On','Ranked By'])
                        # df_no_rank.to_excel(writer, sheet_name='Ranking', index=False)

                        else:
                            rank_calc.append(df_rank_temp['@calculation'].iloc[0])
                            if '@top' in df_rank_temp.columns.values:
                                rank_top.append(df_rank_temp['@top'].iloc[0])
                                rank_bottom.append('')
                            if '@bottom' in df_rank_temp.columns.values:
                                rank_bottom.append(df_rank_temp['@bottom'].iloc[0])
                                rank_top.append('')
                            if 'rankedBy' in df_rank_temp.columns.values:
                                rank_rankedby.append(
                                    df_rank_temp['rankedBy'].iloc[0].replace('=', '').replace('[', '').replace(']', ''))
                            else:
                                rank_rankedby.append('')
                            rank_based.append(df_rank_temp['basedOn'].iloc[0].replace('=', ''))
                            rank_report_id.append(rank_id_list[i])
                            rank_eid.append(rank_id_list2[j])
                            rank_flag = 0
                            for index in def_element_final.index:
                                if def_element_final['EID'][index] == rank_id_list2[j] and \
                                        def_element_final['Report ID'][
                                            index] == rank_id_list[i]:
                                    rank_req_id.append(def_element_final['Requirement ID'][index])
                                    rank_flag = 1
                                    break
                            if rank_flag == 0:
                                rank_req_id.append('')
                            flag = 0
                            for index in def_element_final.index:
                                if def_element_final['EID'][index] == rank_id_list2[j] and \
                                        def_element_final['Report ID'][
                                            index] == rank_id_list[i] and def_element_final['Query Name'][index] == \
                                        df_rank_temp['basedOn'].iloc[0].replace('=', ''):
                                    rank_dp.append(def_element_final['DataProvider Name'][index])
                                    rank_flag = 1
                                    break
                                else:
                                    rank_dp.append('')
                                    rank_flag = 1
                                    break

                            if rank_flag == 0:
                                rank_dp.append('')
        for i in range(len(rank_based)):
            # print('rank_based:',rank_based[i])
            if '].[' in rank_based[i]:
                query = rank_based[i].split('].[')[1].split(']')[0]
                rank_based[i] = query
            else:
                rank_based[i] = rank_based[i].replace('[', '').replace(']', '')
        df_rank['Requirement ID'] = rank_req_id
        df_rank['Report ID'] = rank_report_id
        df_rank['EID'] = rank_eid
        df_rank['Calculation'] = rank_calc
        df_rank['Top'] = rank_top
        df_rank['Bottom'] = rank_bottom
        df_rank['Based On'] = rank_based
        df_rank['Ranked By'] = rank_rankedby
        df_rank['DataProvider Name'] = rank_dp
        df_rank.to_excel(writer, sheet_name='Ranking', index=False)

                # Add If condition to check rank and uncomment 1023,1024. goto line 1021 PASS condition

        # df_no_rank = pd.DataFrame(columns=['Requirement ID', 'Report ID', 'EID', 'Calculation', 'Top', 'Bottom', 'DataProvider Name','Based On','Ranked By'])

        # df_no_rank.to_excel(writer, sheet_name='Ranking', index=False)

        ############################################################################# Ranking Ends ############################################################################

        ############################################################################# InputControl Start ############################################################################

        response = json.loads(get_response(doc_list, 'inputcontrols', 0, int_link)[0])
        
        print( "##Get ALL InputControl",response)
        response = OrderedDict(response)
        df = pd.DataFrame.from_dict(response['inputcontrols']['inputcontrol'])
        df1 = df
        print('rep_list:', rep_list)
        for i in rep_list.keys():
            link = get_response(doc_list, 'reports', 0, int_link)[1] + "/" + str(i) + "/inputcontrols"
            response = requests.request("GET", link, headers=headers, data=payload)
            response = json.loads(response.text)
            response = OrderedDict(response)
            df = pd.DataFrame.from_dict(response['inputcontrols']['inputcontrol'])
            df1 = df1.append(df, ignore_index=True)
        # print('df1:',df1)

        # Check if Input controls are present :

        if len(df1.index) != 0:

            print('df1[id]:',df1['id'])
            df = def_element_final[['Report ID', 'Report Name']].drop_duplicates()
            df['Report ID'] = df['Report ID'].astype(str)
                                
            df1['rep_var'] = df1['id'].str.split('.').str[0]
                            
            df1['Report ID'] = py.where(df1['rep_var'].astype(str) == 'D', 'Document',
                                        df1['rep_var'].replace(regex=r'^[A-Z,a-z]', value=''))
            df1 = pd.merge(df1, df, on=['Report ID'], how='left')
            df1['Report/Dashboard Name'] = py.where(df1['Report ID'].astype(str) == 'Document', rep_name[0],
                                                    df1['Report Name'])
            # df1['Report/Dashboard Name'] = py.where(df1['Report ID'].astype(str) == 'Document', "Sales & Margin Report",
            #                                         df1['Report Name'])
            df1['Affected'] = py.where(df1['Report ID'].astype(str) == 'Document', 'Across Report', 'Specic Dashboard')
            df1['Filter Name'] = df1['name']
            df = df1[['id', 'Affected', 'Report/Dashboard Name', 'Filter Name']]

            allowNullValueSelection = []
            allowAllValuesSelection = []
            numberOfLines = []
            selecttype = []
            selection = []

            ##Get Input controls details
            size = df1.shape[0]
            ic_list = df1['id'].tolist()
            df_ic_detail = pd.DataFrame(index=range(0, size),
                                        columns=['id', 'Operation', 'dpid', 'grouping', 'default', 'custom'])
            for item in ic_list:
                if item.split('.')[0] == 'D':
                    link = int_link + doc_list[0] + '/inputcontrols?allInfo=true'
                else:
                    link = int_link + doc_list[0] + '/reports/' + item.split('.')[0][1:] + '/inputcontrols?allInfo=true'
                response = requests.request("GET", link, headers=headers, data=payload)
                response = OrderedDict((json.loads(response.text)))
                df = pd.DataFrame.from_dict(response['inputcontrols']['inputcontrol'])
                for ind in df.index:
                    if df['id'][ind] == item:
                        if df['selection'][ind].get('@all') == 'true':
                            selection.append('ALL')
                        else:
                            selection.append(','.join(df['selection'][ind].get('value')))

            
            i = 0
            for item in ic_list:
                if item.split('.')[0] == 'D':
                    link = int_link + doc_list[0] + '/inputcontrols/' + item

                else:
                    link = int_link + doc_list[0] + '/reports/' + item.split('.')[0][1:] + '/inputcontrols/' + item
                response = requests.request("GET", link, headers=headers, data=payload)
                response = OrderedDict((json.loads(response.text)))
                df = pd.DataFrame.from_dict(response['inputcontrol'])
                if 'default' in df.index and 'custom' in df.index:
                    df = df.loc[['@refId', '@operator', 'default', '@eligibility', 'custom', '@allowNullValueSelection',
                                '@allowAllValuesSelection'], :]
                    for index in df.loc['default'].index:
                        if isinstance(df.loc['default'][index], dict):
                            df_ic_detail['default'].iloc[i] = df.loc['default'][index]['value']
                    for index in df.loc['custom'].index:
                        if isinstance(df.loc['custom'][index], dict):
                            df_ic_detail['custom'].iloc[i] = df.loc['custom'][index]['value']
                    df_ic_detail['Operation'].iloc[i] = df.loc['@operator'][2]
                elif (('default' not in df.index) and ('custom' in df.index)):
                    df = df.loc[['@refId', '@operator', '@eligibility', 'custom'], :]
                    df_ic_detail['Operation'].iloc[i] = df.loc['@operator'][2]
                    for index in df.loc['custom'].index:
                        if isinstance(df.loc['custom'][index], dict):
                            df_ic_detail['custom'].iloc[i] = df.loc['custom'][index]['value']
                elif (('default' in df.index) and ('custom' not in df.index)):
                    df = df.loc[['@refId', '@operator', 'default', '@eligibility'], :]
                    df_ic_detail['Operation'].iloc[i] = df.loc['@operator'][2]
                    for index in df.loc['default'].index:
                        if isinstance(df.loc['default'][index], dict):
                            df_ic_detail['default'].iloc[i] = df.loc['default'][index]['value']
                else:
                    df = df.reindex(['@refId', '@operator', '@eligibility'])
                    df_ic_detail['Operation'].iloc[i] = df.loc['@operator'][2]
                response = requests.request("GET", link, headers=headers, data=payload)
                response = json.loads(response.text, object_pairs_hook=OrderedDict)
                df2 = pd.DataFrame.from_dict(response['inputcontrol'])
                diffList = (list(set(['@allowNullValueSelection', '@allowAllValuesSelection', '@numberOfLines']).difference(
                    df2.index.values)))
                for j in diffList:
                    emptydataframe = (pd.Series(name=j))
                if len(diffList) > 0:
                    df2 = (df2.append(emptydataframe))
                df2 = df2.reindex(['@allowNullValueSelection', '@allowAllValuesSelection', '@numberOfLines'])
                selecttype.append(df2.columns[2])
                allowNullValueSelection.append(df2.iloc[0, 2])
                allowAllValuesSelection.append(df2.iloc[1, 2])
                numberOfLines.append(df2.iloc[2, 2])
                print("df",df)
                df_ic_detail['id'].iloc[i] = df.loc['@refId']['id']
                if 'assignedDataObject' in df.columns:
                    df_ic_detail['dpid'].iloc[i] = df.loc['@refId']['assignedDataObject']
                if 'assignedDataObjects' in df.columns:
                    df_ic_detail['dpid'].iloc[i] = df.loc['@refId']['assignedDataObjects']
                df_ic_detail['grouping'].iloc[i] = df.loc['@eligibility']['groupingInfo']
                i = i + 1

            ##Joining Of dataframes to get the dataprovider/variable/mergedimension column name

            df_inputcontrol = df1
            df_inputcontrol = pd.merge(df_inputcontrol, df_ic_detail, how='inner', left_on=['id'], right_on=['id'])
            df_dp_final.rename(columns={'name': 'Calculated Field Name'}, inplace=True)
            df_dp_name = pd.concat([df_var[['id', 'Calculated Field Name']], df_dp_final[['id', 'Calculated Field Name']],
                                    df_link[['id', 'Calculated Field Name']]], ignore_index=True)
            df_dp_name.rename(columns={'id': 'dpid'}, inplace=True)
            df_inputcontrol = pd.merge(df_inputcontrol, df_dp_name, how='inner', left_on=['dpid'], right_on=['dpid'])
            df_inputcontrol.rename(columns={'Calculated Field Name': 'Filter Column'}, inplace=True)
            df_inputcontrol = df_inputcontrol[
                ['id', 'Affected', 'Report/Dashboard Name', 'Filter Name', 'Filter Column', 'Operation', 'default',
                'custom',
                'grouping']]
            for index in df_inputcontrol.index:
                if "[u'" in str(df_inputcontrol['default'][index]) or ", u'" in str(df_inputcontrol['default'][index]):
                    df_inputcontrol['default'][index] = str(df_inputcontrol['default'][index]).replace("[u'", "['").replace(
                        ", u'", ", '").replace('[', '').replace(']', '')
                elif str(df_inputcontrol['default'][index]) == 'nan':
                    df_inputcontrol['default'][index] = ''
                else:
                    df_inputcontrol['default'][index] = str(df_inputcontrol['default'][index]).replace('[', '').replace(']',
                                                                                                                        '')
            for index in df_inputcontrol.index:
                if "[u'" in str(df_inputcontrol['custom'][index]) or ", u'" in str(df_inputcontrol['custom'][index]):
                    df_inputcontrol['custom'][index] = str(df_inputcontrol['custom'][index]).replace("[u'", "['").replace(
                        ", u'", ", '").replace('[', '').replace(']', '')
                elif str(df_inputcontrol['custom'][index]) == 'nan':
                    df_inputcontrol['custom'][index] = ''
                else:
                    df_inputcontrol['custom'][index] = str(df_inputcontrol['custom'][index]).replace('[', '').replace(']',
                                                                                                                    '')
            df_inputcontrol['allow Null Value Selection'] = pd.Series(allowNullValueSelection)
            df_inputcontrol['allow All Values Selection'] = pd.Series(allowAllValuesSelection)
            df_inputcontrol['number Of Lines'] = pd.Series(numberOfLines)
            df_inputcontrol['select type'] = pd.Series(selecttype)
            df_inputcontrol['selection'] = pd.Series(selection)
            for i, rows in df_inputcontrol.iterrows():
                if rows['id'].startswith('D'):
                    df_inputcontrol.loc[i, 'Report/Dashboard Name'] = 'All elements'
            df_inputcontrol.to_excel(writer, sheet_name='InputControls', index=False)

        else:
            print("Creating empty sheet for Input Controls")

            df_no_input_control_elems = pd.DataFrame(columns=['id', 'Affected', 'Report/Dashboard Name', 'Filter Name', 'Filter Column', 'Operation', 'default','custom','grouping'])

            df_no_input_control_elems.to_excel(writer, sheet_name='InputControls', index=False)

        #######################################################################   Input Controls Ends #######################################################################
        print("Come out of Input Controls")

        #######################################################################   Conditional Formatting Start #######################################################################

        df_conditional = pd.DataFrame(
            columns=['Alerter Name', 'Object Name', 'Operator', 'Value', 'Font Format', 'Background Format'])
        con_an = []
        con_on = []
        con_op = []
        con_val = []
        con_fo = []
        con_bg = []
        ruleid = []
        for id in doc_list:

            conditionlink = int_link + id + '/alerters'
            condition_response = requests.request("GET", conditionlink, headers=headers, data=payload)
            condition_response = OrderedDict(json.loads(condition_response.text))
            df_condition_temp = pd.DataFrame.from_dict(condition_response['alerters']['alerter'])
            if 'id' in df_condition_temp:
                id_list = df_condition_temp['id']
                name_list = df_condition_temp['name']
            else:
                id_list = []
                name_list = []
            for lst_id in range(len(id_list)):
                i = lst_id
                lst_id = str(id_list[lst_id])
                cond_details_link = conditionlink + '/' + lst_id
                condition_response_final = requests.request("GET", cond_details_link, headers=headers, data=payload)
                condition_response_final = OrderedDict(json.loads(condition_response_final.text))
                df_condition_final = pd.DataFrame.from_dict(condition_response_final['alerter']['rule'])
                for index, row in df_condition_final.iterrows():
                    ruleid.append(row['id'])
                    my_list = row['conditions']['condition']
                    my_list = json.dumps(my_list)
                    my_list = ast.literal_eval(my_list)
                    for di in my_list:
                        con_val.append(','.join(di.get('operand')))
                        con_op.append(di.get('@operator'))
                        con_on.append(di.get('@expressionId'))
                        con_an.append(name_list[i])
                    my_list = row['action']['style']
                    my_list = json.dumps(my_list)
                    my_list = ast.literal_eval(my_list)
                    # con_bg.append(my_list.get('background').get('color'))
                    # print type(my_list.get('background').get('color'))
                    if my_list.get('background').get('color'):
                        con_bg.append(
                            json.dumps(my_list.get('background').get('color')).split('{"@rgb": "')[1].split('"}')[0])
                    else:
                        con_bg.append('')
                    con_fo.append(my_list.get('font').get('@rgb'))
        dp_id_list = df_dp_final_bkp1["id"].tolist()
        dp_name_list = df_dp_final_bkp1["name"].tolist()
        for i in range(len(con_on)):
            for j in range(len(dp_id_list)):
                if dp_id_list[j] == con_on[i]:
                    con_on[i] = dp_name_list[j]
                    break
        df_conditional['Alerter Name'] = con_an
        df_conditional['Object Name'] = con_on
        df_conditional['Operator'] = con_op
        df_conditional['Value'] = con_val
        df_conditional['Font Format'] = con_fo
        df_conditional['Background Format'] = con_bg
        df_conditional['Rule id'] = ruleid

        df_conditional.to_excel(writer, sheet_name='Conditional Formatting', index=False)

        # Add If condition to check conditional foramtiing and uncomment below 2 lines.

        # df_no_conditional_formatting = pd.DataFrame(columns=['Alerter Name', 'Object Name', 'Operator', 'Value', 'Font Format', 'Background Format', 'Rule id'])

        # df_no_conditional_formatting.to_excel(writer, sheet_name='Conditional Formatting', index=False)


        #######################################################################   Conditional Formatting End #######################################################################

        # removing filter column and duplicate rows

        def_element_final2 = def_element_final.drop('filter', axis=1)
        def_element_final3 = def_element_final2[
            ['Requirement ID', 'EID', 'Report ID', 'Report Name', 'Chart Title', 'Chart Type', 'DataProvider Name',
             'Query Name', 'hyper link', 'Field Name', 'IsMerge', 'Qualification', 'Aggregation']].drop_duplicates()
        def_element_final3.to_excel(writer, sheet_name='Worksheets', index=False)

        # adding formulas to calculation sheet
        cal_for = [[]]
        i = 0
        flag = 0
        for index, row in def_element_final3.iterrows():
            if row['Field Name'].startswith('formula') and row['Field Name'].replace('formula', '').isdigit():
                cal_for.append(
                    ['', row['Field Name'], row['DataProvider Name'], row['Query Name'], row['Qualification']])
                flag = 1
            i = i + 1

        if flag == 1:
            df_for = pd.DataFrame(cal_for,
                                  columns=['id', 'Calculated Field Name', 'DataProvider', 'Formula', 'Qualification'])
            df_for = df_for.iloc[1:]
            df_var = df_var.drop('IsMerge', axis=1)
            df_var = pd.concat([df_var, df_for])
            df_var.to_excel(writer, sheet_name='Calculations', index=False)

        writer.save()

        print('COMPLETED SUCCESSFULLY !')

        # Generating SQL Files

        print('Generating SQL Files......')

        cols = [0, 4, 6]
        df_excel = pd.read_excel(output_excel, sheet_name='DataProvider', usecols=cols)
        df_excel = df_excel.drop_duplicates(subset=['pid', 'query', 'DataProvider Name'])

        for index, row in df_excel.iterrows():
            if str(row['query']) != 'nan':
                f = open((output_excel[:-(len("_BO_DFT.xlsx"))]).split('\\')[-1] + "_" + row['pid'] + "_" + row[
                    'DataProvider Name'] + ".sql", 'w')
                f.write(str(row['query']))
                f.close()

        # GENERATING TSD#

        def cell_style(cell):
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='top', shrink_to_fit=True, wrapText=True)

        excelTSD_path = "BoT TSD Template.xlsx"
        excelTSD_wb = openpyxl.load_workbook(excelTSD_path)

        excelExtract_path = output_excel
        reportName = (excelExtract_path[:-(len("_BO_DFT.xlsx"))]).split('\\')[-1]
        print("Generating TSD Excel for : " + reportName)

        # UPDATING COVER SHEET TAB
        excelTSD_ws = excelTSD_wb["Cover Sheet"]
        excelTSD_ws['B5'] = (reportName[::-1].split('/')[0])[::-1]
        excelTSD_ws['B6'] = RepID

        # UPDATING REPORT QUERY TAB
        excelTSD_ws = excelTSD_wb["Report Query"]
        cols = [2, 4, 11, 3, 6]
        df_reportQuery = pd.read_excel(excelExtract_path, sheet_name='DataProvider', usecols=cols)
        df_reportQuery = df_reportQuery.drop_duplicates(
            subset=['dataSourceType', 'File Name', 'DataProvider Name', 'query', 'Custom Query Indicator'])
        column_counter = 3
        for index, row in df_reportQuery.iterrows():
            excelTSD_ws.cell(row=4, column=column_counter).value = row["DataProvider Name"].replace('_', ' ')
            excelTSD_ws.cell(row=5, column=column_counter).value = row["dataSourceType"].title()
            if row["Custom Query Indicator"] == '' or pd.isnull(row["Custom Query Indicator"]):
                excelTSD_ws.cell(row=6, column=column_counter).value = 'False'
            else:
                excelTSD_ws.cell(row=6, column=column_counter).value = str(row["Custom Query Indicator"]).title()
            excelTSD_ws.cell(row=7, column=column_counter).value = row["File Name"]
            excelTSD_ws.cell(row=8, column=column_counter).value = row["query"]
            cell_style(excelTSD_ws.cell(row=4, column=column_counter))
            cell_style(excelTSD_ws.cell(row=5, column=column_counter))
            cell_style(excelTSD_ws.cell(row=6, column=column_counter))
            cell_style(excelTSD_ws.cell(row=7, column=column_counter))
            cell_style(excelTSD_ws.cell(row=8, column=column_counter))
            column_counter = column_counter + 1

        # Updating Report Details tab
        # Updating report variables part Report Details tab
        excelTSD_ws = excelTSD_wb["Report Details"]
        col_filter = [1, 3, 4, 2]
        df_calculatedQuery = pd.read_excel(excelExtract_path, sheet_name='Calculations', usecols=col_filter)
        df_calculatedQuery = df_calculatedQuery.drop_duplicates(
            subset=['Calculated Field Name', 'Formula', 'Qualification', 'DataProvider'])
        row_counter = 6
        for i, rows in df_calculatedQuery.iterrows():
            excelTSD_ws.cell(row=row_counter, column=1).value = rows['Calculated Field Name']
            excelTSD_ws.cell(row=row_counter, column=2).value = rows['Formula']
            excelTSD_ws.cell(row=row_counter, column=3).value = rows['Qualification']
            excelTSD_ws.cell(row=row_counter, column=4).value = rows['DataProvider']
            cell_style(excelTSD_ws.cell(row=row_counter, column=1))
            cell_style(
                excelTSD_ws.cell(row=row_counter, column=2))  # .alignment=Alignment(horizontal="left", vertical="top")
            cell_style(
                excelTSD_ws.cell(row=row_counter, column=3))  # .alignment=Alignment(horizontal="left", vertical="top")
            cell_style(
                excelTSD_ws.cell(row=row_counter, column=4))  # .alignment=Alignment(horizontal="left", vertical="top")
            row_counter = row_counter + 1

        # update merge dimensions part in report details tab
        excelTSD_ws = excelTSD_wb["Report Details"]
        col_filter = [1, 7, 2]
        df_calculatedQuery = pd.read_excel(excelExtract_path, sheet_name='Merge_Dim', usecols=col_filter)
        df_calculatedQuery = df_calculatedQuery.drop_duplicates(
            subset=['Calculated Field Name', 'MergeDimension', '@dataType'])
        row_counter = 6
        for i, rows in df_calculatedQuery.iterrows():
            primary_source_query = rows['MergeDimension'].split('[')[1].split('].[')[-1]
            primary_source_query = primary_source_query.replace("].", '')
            excelTSD_ws.cell(row=row_counter, column=6).value = rows['Calculated Field Name']
            excelTSD_ws.cell(row=row_counter, column=7).value = rows['MergeDimension']
            excelTSD_ws.cell(row=row_counter, column=8).value = rows['@dataType']
            excelTSD_ws.cell(row=row_counter, column=9).value = primary_source_query
            cell_style(excelTSD_ws.cell(row=row_counter, column=6))
            cell_style(excelTSD_ws.cell(row=row_counter, column=7))
            cell_style(excelTSD_ws.cell(row=row_counter, column=8))
            cell_style(excelTSD_ws.cell(row=row_counter, column=9))
            row_counter = row_counter + 1

        # updating report filters tab
        # updating report filters part report filters tab
        excelTSD_ws = excelTSD_wb["Report Filters"]
        col_filter = [2, 3, 4, 5, 6, 7]
        df_filtertQuery = pd.read_excel(excelExtract_path, sheet_name='FilterDetails', usecols=col_filter)
        df_filtertQuery = df_filtertQuery.drop_duplicates(
            subset=['Chart Title', 'Report Name', 'filter', 'DataProvider Name', 'operation', 'value'])
        df_filtertQuery = df_filtertQuery.sort_values('Chart Title')
        row_counter = 6
        for i, rows in df_filtertQuery.iterrows():
            excelTSD_ws.cell(row=row_counter, column=1).value = rows['Report Name']
            excelTSD_ws.cell(row=row_counter, column=2).value = rows['Chart Title']
            excelTSD_ws.cell(row=row_counter, column=3).value = rows['filter']
            excelTSD_ws.cell(row=row_counter, column=4).value = rows['DataProvider Name']
            excelTSD_ws.cell(row=row_counter, column=5).value = rows['operation']
            excelTSD_ws.cell(row=row_counter, column=6).value = rows['value']
            cell_style(excelTSD_ws.cell(row=row_counter, column=1))
            cell_style(excelTSD_ws.cell(row=row_counter, column=2))
            cell_style(excelTSD_ws.cell(row=row_counter, column=3))
            cell_style(excelTSD_ws.cell(row=row_counter, column=4))
            cell_style(excelTSD_ws.cell(row=row_counter, column=5))
            cell_style(excelTSD_ws.cell(row=row_counter, column=6))
            row_counter = row_counter + 1
        # input controls
        excelTSD_ws = excelTSD_wb["Report Filters"]
        col_input = [0, 1, 3, 2, 4, 7, 8, 6, 9, 10, 12]
        df_inputControls = pd.read_excel(excelExtract_path, sheet_name='InputControls', usecols=col_input)
        df_inputControls = df_inputControls.sort_values('Report/Dashboard Name')
        row_counter = 6
        col_input = [2, 6]
        df_ws = pd.read_excel(excelExtract_path, sheet_name='Worksheets', usecols=col_input)
        for i, rows in df_inputControls.iterrows():
            if rows['id'].startswith('R'):
                excelTSD_ws.cell(row=row_counter, column=8).value = 'Report'
            if rows['id'].startswith('D'):
                excelTSD_ws.cell(row=row_counter, column=8).value = 'Document'
            excelTSD_ws.cell(row=row_counter, column=9).value = rows['Report/Dashboard Name']
            excelTSD_ws.cell(row=row_counter, column=10).value = rows['Filter Column']
            for j, rows2 in df_ws.iterrows():
                if str(rows2['Report ID']) == str(rows['id'][1:].split('.')[0]):
                    excelTSD_ws.cell(row=row_counter, column=11).value = rows2['DataProvider Name']
                    break
            link = int_link + doc_list[0] + '/inputcontrols/?allInfo=true'
            print(link)
            response = requests.request("GET", link, headers=headers, data=payload)
            response = OrderedDict((json.loads(response.text)))
            df_ic = pd.DataFrame.from_dict(response['inputcontrols'])
            print(df_ic)
            excelTSD_ws.cell(row=row_counter, column=12).value = rows['select type']
            excelTSD_ws.cell(row=row_counter, column=13).value = rows['allow All Values Selection']
            excelTSD_ws.cell(row=row_counter, column=14).value = rows['allow Null Value Selection']
            excelTSD_ws.cell(row=row_counter, column=15).value = rows['default']
            excelTSD_ws.cell(row=row_counter, column=16).value = rows['custom']
            cell_style(excelTSD_ws.cell(row=row_counter, column=8))
            cell_style(excelTSD_ws.cell(row=row_counter, column=9))
            cell_style(excelTSD_ws.cell(row=row_counter, column=10))
            cell_style(excelTSD_ws.cell(row=row_counter, column=11))
            cell_style(excelTSD_ws.cell(row=row_counter, column=12))
            cell_style(excelTSD_ws.cell(row=row_counter, column=13))
            cell_style(excelTSD_ws.cell(row=row_counter, column=14))
            cell_style(excelTSD_ws.cell(row=row_counter, column=15))
            cell_style(excelTSD_ws.cell(row=row_counter, column=16))
            row_counter = row_counter + 1

        # driller report
        excelTSD_ws = excelTSD_wb["Report Filters"]
        col_driller = [0, 1, 2, 3, 4, 5]
        df_drillerQuery = pd.read_excel(excelExtract_path, sheet_name='DrillerDetails', usecols=col_driller)
        df_drillerQuery = df_drillerQuery.drop_duplicates(
            subset=['Report ID', 'Report Name', 'name', 'Qualification', 'value'])
        row_counter = 6
        col_input = [2, 6]
        df_ws = pd.read_excel(excelExtract_path, sheet_name='Worksheets', usecols=col_input)
        for i, rows in df_drillerQuery.iterrows():
            excelTSD_ws.cell(row=row_counter, column=18).value = rows['Report Name']
            excelTSD_ws.cell(row=row_counter, column=19).value = rows['name']
            for j, rows2 in df_ws.iterrows():
                if rows2['Report ID'] == rows['Report ID']:
                    excelTSD_ws.cell(row=row_counter, column=20).value = rows2['DataProvider Name']
                    break
            excelTSD_ws.cell(row=row_counter, column=21).value = rows['Qualification']
            excelTSD_ws.cell(row=row_counter, column=22).value = rows['value']
            cell_style(excelTSD_ws.cell(row=row_counter, column=18))
            cell_style(excelTSD_ws.cell(row=row_counter, column=19))
            cell_style(excelTSD_ws.cell(row=row_counter, column=20))
            cell_style(excelTSD_ws.cell(row=row_counter, column=21))
            cell_style(excelTSD_ws.cell(row=row_counter, column=22))
            row_counter = row_counter + 1

        # Updating Ranking & Alerters tab
        # updating ranking tab in Ranking & Alerters

        excelTSD_ws = excelTSD_wb["Ranking & Alerters"]
        col_filter = [2, 4, 5, 8, 7, 3]
        df_calculatedQuery = pd.read_excel(excelExtract_path, sheet_name='Ranking', usecols=col_filter)
        df_calculatedQuery = df_calculatedQuery.drop_duplicates(
            subset=['EID', 'Top', 'Bottom', 'Ranked By', 'Based On', 'Calculation'])
        row_counter = 6
        for i, rows in df_calculatedQuery.iterrows():
            excelTSD_ws.cell(row=row_counter, column=1).value = rows['EID']
            excelTSD_ws.cell(row=row_counter, column=2).value = rows['Top']
            excelTSD_ws.cell(row=row_counter, column=3).value = rows['Bottom']
            excelTSD_ws.cell(row=row_counter, column=4).value = rows['Ranked By']
            excelTSD_ws.cell(row=row_counter, column=5).value = rows['Based On']
            excelTSD_ws.cell(row=row_counter, column=6).value = rows['Calculation']
            cell_style(excelTSD_ws.cell(row=row_counter, column=1))
            cell_style(excelTSD_ws.cell(row=row_counter, column=2))
            cell_style(excelTSD_ws.cell(row=row_counter, column=3))
            cell_style(excelTSD_ws.cell(row=row_counter, column=4))
            cell_style(excelTSD_ws.cell(row=row_counter, column=5))
            cell_style(excelTSD_ws.cell(row=row_counter, column=6))
            row_counter = row_counter + 1

        col_filter = [0, 1, 2, 3, 4, 5]
        df_con = pd.read_excel(excelExtract_path, sheet_name='Conditional Formatting', usecols=col_filter)
        row_counter = 6
        for i, rows in df_con.iterrows():
            excelTSD_ws.cell(row=row_counter, column=8).value = rows['Alerter Name']
            if rows["Object Name"] == '' or pd.isnull(rows["Object Name"]):
                excelTSD_ws.cell(row=row_counter, column=9).value = 'Cell content'
            else:
                excelTSD_ws.cell(row=row_counter, column=9).value = rows['Object Name']
            excelTSD_ws.cell(row=row_counter, column=10).value = rows['Operator']
            excelTSD_ws.cell(row=row_counter, column=11).value = rows['Value']
            excelTSD_ws.cell(row=row_counter, column=12).value = rows['Font Format']
            excelTSD_ws.cell(row=row_counter, column=13).value = rows['Background Format']
            cell_style(excelTSD_ws.cell(row=row_counter, column=8))
            cell_style(excelTSD_ws.cell(row=row_counter, column=9))
            cell_style(excelTSD_ws.cell(row=row_counter, column=10))
            cell_style(excelTSD_ws.cell(row=row_counter, column=11))
            cell_style(excelTSD_ws.cell(row=row_counter, column=12))
            cell_style(excelTSD_ws.cell(row=row_counter, column=13))
            row_counter = row_counter + 1

        # Updating Element Mapping Sheet
        excelTSD_ws = excelTSD_wb["Element Mapping"]
        col_filter = [3, 4, 5, 9, 12, 11, 10, 6, 7]
        df_worksheets = pd.read_excel(excelExtract_path, sheet_name='Worksheets', usecols=col_filter)
        df_worksheets = df_worksheets.drop_duplicates(
            subset=['Report Name', 'Chart Title', 'Chart Type', 'Field Name', 'Aggregation', 'Qualification', 'IsMerge',
                    'DataProvider Name', 'Query Name'])
        cols = [4, 12]
        df_dp = pd.read_excel(excelExtract_path, sheet_name='DataProvider', usecols=cols)
        df_dp = df_dp.drop_duplicates(subset=['DataProvider Name', 'Universe Name'])
        row_counter = 5
        for i, rows in df_worksheets.iterrows():
            excelTSD_ws.cell(row=row_counter, column=1).value = rows['Report Name']
            excelTSD_ws.cell(row=row_counter, column=2).value = rows['Chart Title']
            excelTSD_ws.cell(row=row_counter, column=3).value = rows['Chart Type']
            excelTSD_ws.cell(row=row_counter, column=4).value = rows['Query Name']
            excelTSD_ws.cell(row=row_counter, column=5).value = rows['Field Name']
            excelTSD_ws.cell(row=row_counter, column=6).value = rows['Aggregation']
            excelTSD_ws.cell(row=row_counter, column=7).value = rows['Qualification']
            excelTSD_ws.cell(row=row_counter, column=8).value = rows['IsMerge']
            excelTSD_ws.cell(row=row_counter, column=9).value = rows['DataProvider Name']
            for j, rows2 in df_dp.iterrows():
                if str(rows2['DataProvider Name']) == str(rows['DataProvider Name']):
                    excelTSD_ws.cell(row=row_counter, column=10).value = rows2['Universe Name']
                    break
            cell_style(excelTSD_ws.cell(row=row_counter, column=1))
            cell_style(excelTSD_ws.cell(row=row_counter, column=2))
            cell_style(excelTSD_ws.cell(row=row_counter, column=3))
            cell_style(excelTSD_ws.cell(row=row_counter, column=4))
            cell_style(excelTSD_ws.cell(row=row_counter, column=5))
            cell_style(excelTSD_ws.cell(row=row_counter, column=6))
            cell_style(excelTSD_ws.cell(row=row_counter, column=7))
            cell_style(excelTSD_ws.cell(row=row_counter, column=8))
            cell_style(excelTSD_ws.cell(row=row_counter, column=9))
            cell_style(excelTSD_ws.cell(row=row_counter, column=10))
            row_counter = row_counter + 1

        excelTSD_wb.save(reportName + '_TSD.xlsx')
        excelTSD_wb.close()

        response = {
            'RepID': RepID,
            'status': 'Successful',
            'file_name': reportName + '_TSD.xlsx'}

        print("Response", response)

    except IndexError as ie:
        print("Error : {0} for RepID : {1}".format(ie, RepID))
        response = {
            'RepID': RepID,
            'status': 'Unsuccessful'}

        print("Response", response)

    return response


#Universe Extraction Functions begin
#### Parameterize Auth Type - secEnterprise

def GetPayload(login_link, UserID, Password, Auth):

    payload = '<attrs xmlns="' + login_link + '"><attr name="userName" type="string">' + UserID + '</attr><attr name="password" type="string">' + Password + '</attr><attr name="auth" type="string" possibilities="secEnterprise,secLDAP,secWinAD,secSAPR3">' + Auth + '</attr></attrs>'
    return (payload)


# def GetLoginToken_unv(login_link, UserID, Password):
#     url = 'http://cvyhj1a18:6405/biprws/logon/long'
#     payload = GetPayload(login_link, UserID, Password)
#     headers = {
#         'content-type': 'application/xml',
#         'accept': 'application/xml'
#     }
#     try:
#         response = requests.request("POST", url, headers=headers, data=payload)
#         if response.status_code == 200:
#             Status = 'Success'
#             return (response.headers['X-SAP-LogonToken'], Status)
#         else:
#             # print("Failed to generate token")
#             Status = 'Failed'
#             return (1, Status)
#     except:
#         return (2, 'Connection Error')


# def getJSON(UserID, Password, LoginToken, FolderID, login_link):
#     doc_link = 'http://cvyhj1a18:6405/biprws/raylight/v1/universes/' + FolderID
#     headers = {
#         'content-type': 'application/x-www-form-urlencoded',
#         'accept': 'application/json',
#         'X-SAP-LogonToken': LoginToken
#     }
#     payload = GetPayload(login_link, UserID, Password)
#     response = requests.request("GET", doc_link, headers=headers, data=payload)
#     # print(response.status_code)
#     return (response.json())


# def items(value):
#     if 'id' in value:
#         TempObjectID = value['id']
#     else:
#         TempObjectID = ''

#     if 'name' in value:
#         TempObjectName = value['name']
#     else:
#         TempObjectName = ''

#     if 'description' in value:
#         TempObjectDec = value['description']
#     else:
#         TempObjectDec = ''

#     if '@type' in value:
#         TempObjectType = value['@type']
#     else:
#         TempObjectType = ''
#     if 'path' in value:
#         TempPath = value['path']
#     else:
#         TempPath = ''

#     if '@dataType' in value:
#         TempDataType = value['@dataType']
#     else:
#         TempDataType = ''

#     if '@hasLov' in value:
#         TempHasLOV = value['@hasLov']
#     else:
#         TempHasLOV = ''
#     return TempObjectID, TempObjectName, TempObjectDec, TempObjectType, TempPath, TempDataType, TempHasLOV


# def Extraction(ret):
#     FolderName = []
#     ObjectName = []
#     ObjectID = []
#     ObjectType = []
#     ObjectDec = []
#     Path = []
#     DataType = []
#     HasLOV = []
#     BaseAttribName = []
#     BaseAttribType = []
#     for item in ret['universe']['outline']['folder']:
#         # print(i['name'])
#         # if i['name']=='Time period':
#         # print(i)
#         # print(i['name'])
#         if 'folder' in item:
#             for subitem in item['folder']:
#                 for subsubitem in (subitem['item']):
#                     if 'item' in subsubitem:
#                         for subsubsubitem in subsubitem['item']:
#                             TempObjectID, TempObjectName, TempObjectDec, TempObjectType, TempPath, TempDataType, TempHasLOV = items(
#                                 subsubsubitem)
#                             ObjectID.append(TempObjectID)
#                             ObjectName.append(TempObjectName)
#                             ObjectDec.append(TempObjectDec)
#                             ObjectType.append(TempObjectType)
#                             BaseAttribName.append(subsubitem['name'])
#                             BaseAttribType.append(subsubitem['@type'])
#                             FolderName.append(subitem['name'])
#                             Path.append(TempPath)
#                             DataType.append(TempDataType)
#                             HasLOV.append(TempHasLOV)

#                     else:
#                         TempObjectID, TempObjectName, TempObjectDec, TempObjectType, TempPath, TempDataType, TempHasLOV = items(
#                             subsubitem)
#                         ObjectID.append(TempObjectID)
#                         ObjectName.append(TempObjectName)
#                         ObjectDec.append(TempObjectDec)
#                         ObjectType.append(TempObjectType)
#                         BaseAttribName.append('')
#                         BaseAttribType.append('')
#                         FolderName.append(subitem['name'])
#                         Path.append(TempPath)
#                         DataType.append(TempDataType)
#                         HasLOV.append(TempHasLOV)

#         if 'item' in item:
#             for subitem in (item['item']):
#                 # print(j['item'])
#                 if 'item' in subitem:
#                     for subsubitem in subitem['item']:
#                         TempObjectID, TempObjectName, TempObjectDec, TempObjectType, TempPath, TempDataType, TempHasLOV = items(
#                             subsubitem)
#                         ObjectID.append(TempObjectID)
#                         ObjectName.append(TempObjectName)
#                         ObjectDec.append(TempObjectDec)
#                         ObjectType.append(TempObjectType)
#                         BaseAttribName.append(subitem['name'])
#                         BaseAttribType.append(subitem['@type'])
#                         FolderName.append(item['name'])
#                         Path.append(TempPath)
#                         DataType.append(TempDataType)
#                         HasLOV.append(TempHasLOV)



#                 else:
#                     TempObjectID, TempObjectName, TempObjectDec, TempObjectType, TempPath, TempDataType, TempHasLOV = items(
#                         subitem)
#                     ObjectID.append(TempObjectID)
#                     ObjectName.append(TempObjectName)
#                     ObjectDec.append(TempObjectDec)
#                     ObjectType.append(TempObjectType)
#                     BaseAttribName.append('')
#                     BaseAttribType.append('')
#                     FolderName.append(item['name'])
#                     Path.append(TempPath)
#                     DataType.append(TempDataType)
#                     HasLOV.append(TempHasLOV)

#     df = pd.DataFrame(list(
#         zip(ObjectID, ObjectName, ObjectDec, ObjectType, BaseAttribName, BaseAttribType, FolderName, Path, DataType,
#             HasLOV)),
#                       columns=['Object ID', 'Object Name', 'Object Desc', 'Object Type', 'Base Object Name',
#                                'Base Object Type', 'Folder Name', 'Folder Path', 'Data type', 'Has Lov'])
#     return (df)


# def UniverseDetails(ret):
#     UniverseID = []
#     CuID = []
#     UniverseName = []
#     Type = []
#     UniverseDesc = []
#     PathUniverse = []
#     MaxRowsRetrieved = []
#     MaxRetrievalTime = []
#     UniverseID.append(ret['universe']['id'])
#     CuID.append(ret['universe']['cuid'])
#     UniverseName.append(ret['universe']['name'])
#     Type.append(ret['universe']['type'])
#     if 'description' in ret['universe']:
#         UniverseDesc.append(ret['universe']['description'])
#     else:
#         UniverseDesc.append('')

#     PathUniverse.append(ret['universe']['path'])
#     if 'maxRowsRetrieved' in ret['universe']:
#         MaxRowsRetrieved.append(ret['universe']['maxRowsRetrieved'])
#     else:
#         MaxRowsRetrieved.append('')

#     if 'maxRetrievalTime' in ret['universe']:
#         MaxRetrievalTime.append(ret['universe']['maxRetrievalTime'])
#     else:
#         MaxRetrievalTime.append('')

#     df = pd.DataFrame(
#         list(zip(UniverseID, CuID, UniverseName, Type, UniverseDesc, PathUniverse, MaxRowsRetrieved, MaxRetrievalTime)),
#         columns=['Universe ID', 'Universe Cuid', 'Universe Name', 'Universe Type', 'Universe Description',
#                  'Universe Path', 'Universe MaxRowsRetrieved', 'Universe MaxRetrievalTime'])
#     return (df, ret['universe']['name'])


# def CoverSheetUpdate(Dataframe, FileName):
#     Dataframe = Dataframe[['Universe Name', 'Universe Cuid', 'Universe Description', 'Universe Type', 'Universe Path',
#                            'Universe MaxRowsRetrieved', 'Universe MaxRetrievalTime']]
#     CoversheetCellUpdate = ['B5', 'B6', 'B7', 'B8', 'B9', 'B10', 'B11']
#     val = 0
#     for i in Dataframe.columns:
#         # filename=r'C:\Users\ankitkumgupta\Desktop\Asset Development\Universe TSD Template.xlsx'
#         xfile = openpyxl.load_workbook(FileName)
#         sheet = xfile.get_sheet_by_name('Cover Sheet')
#         sheet[CoversheetCellUpdate[val]] = Dataframe.iloc[0][i]
#         sheet[CoversheetCellUpdate[val]].alignment = Alignment(horizontal='left', vertical='top', shrink_to_fit=True,
#                                                                wrapText=True)
#         xfile.save(FileName)
#         # print(DataFrameUniverse.iloc[0][i])
#         val = val + 1


# def ObjectDetailsUpdate(DataFrame, FileName):
#     wb = load_workbook(FileName)
#     sheet = wb.get_sheet_by_name('Object Details')
#     ColumnCounter = 1
#     for i in DataFrame.columns:
#         RowCounter = 5
#         for j in range(0, len(DataFrame)):
#             # print(j)
#             sheet.cell(row=RowCounter, column=ColumnCounter).value = DataFrame.iloc[j][i]
#             sheet.cell(row=RowCounter, column=ColumnCounter).alignment = Alignment(horizontal='left', vertical='top',
#                                                                                    shrink_to_fit=True, wrapText=True)
#             # print(DataFrame.iloc[j][i])
#             # print(sheet.cell(row=RowCounter, column=ColumnCounter).value)
#             RowCounter = RowCounter + 1
#         ColumnCounter = ColumnCounter + 1
#     wb.save(FileName)


# def WordClean(sentence):
#     sentence = sentence + "\\"
#     while sentence.find('|') and sentence.find('|') > 0:
#         if sentence.find('|'):
#             sentence = sentence.replace(sentence[sentence.find('|'):sentence[sentence.find('|'):].find('\\') + len(
#                 sentence[:sentence.find('|')])], "")
#         # print(sentence)
#     sentence = sentence[:len(sentence) - 1]
#     return (sentence)


# def task(elements, UserID, Password):
#     if elements != '':
#         login_link = 'http://www.sap.com/rws/bip'
#         Token, Status = GetLoginToken_unv(login_link, UserID, Password)
#         if Status == 'Success':
#             LoginToken = Token
#             ret = getJSON(UserID, Password, LoginToken, elements, login_link)
#             if 'error_code' not in ret:
#                 DataFrame = Extraction(ret)
#                 DataFrame['Folder Path'] = DataFrame['Folder Path'].apply(WordClean)
#                 DataFrameUniverse, UniverName = UniverseDetails(ret)
#                 TemplateFilePath = os.path.join(RootPath, "Universe TSD Template.xlsx")
#                 FileName = os.path.join(RootPath, f"Universe Output//{UniverName} Universe TSD.xlsx")
#                 # TemplateFilePath='C://Users//ankitkumgupta//Desktop//Asset Development//Universe TSD Template.xlsx'
#                 # FileName= f"C://Users//ankitkumgupta//Desktop//Asset Development//Universe Output//{UniverName} Universe TSD.xlsx"
#                 shutil.copy2(TemplateFilePath, FileName)
#                 ObjectDetailsUpdate(DataFrame, FileName)
#                 CoverSheetUpdate(DataFrameUniverse, FileName)
#                 return (f"TSD Generated successfully")
#             # print(f"{count}. {UniverName} Universe TSD.xlsx file generated successfully")
#             # DataFrame.to_excel(r'C:\Users\ankitkumgupta\Desktop\Asset Development\SampleOutput.xlsx',index=False)
#             else:
#                 return (f"Invalid Folder ID")
#         elif Status == 'Connection Error':
#             return ("Connection Error")
#         else:
#             return ("Token generation failed")
#     else:
#         return (f"Invalid Folder ID")


# @app.route('/UniverseExtract',methods=['POST'])
# def parallel_unv():
#     l1=[]
#     info = {
#         'username': flask.request.json['username'],
#         'password': flask.request.json['password'],
#         'FolderID': flask.request.json['FolderID']
#     }
#     ListFolderID=info['FolderID'].split(',')
#     #RemoveContents=remove_contents(RootPath+"//Universe Output//")
#     for c in os.listdir(RootPath+"//Universe Output//"):
#         full_path = os.path.join(RootPath+"//Universe Output//", c)
#         if os.path.isfile(full_path):
#             os.remove(full_path)
#             print(full_path+" file deleted")
#     #RemoveContents=remove_contents('C://Users//ankitkumgupta//Desktop//Asset Development//Universe Output//')
#     executor = ProcessPoolExecutor(5)
#     results = [executor.submit(task,elements,info['username'],info['password']) for elements in ListFolderID]
#     while True:
#         count = len([1 for result in results if result.done() == True])
#         print("{} of {} tasks completed".format(count, len(results)))
#         if count == len(results):
#             break
#         time.sleep(2)
#         for i in results:
#             l1.append(i.result())
#             print(i.result())
#         dictionary = dict(zip(ListFolderID, l1))
#     return flask.jsonify(dictionary)


# @app.route('/UniverseDownload/<path:filename>',methods=['GET','POST'])
# def download_unv(filename):
#     DownloadPath=RootPath+"//Universe Output//"
#     return flask.send_from_directory(directory=DownloadPath,filename=filename,as_attachment=True)

# @app.route('/UniverseDownloadAll',methods=['POST'])
# def download_unv_all():
#     Filenames = {
#         'path': flask.request.json['path']
#     }
#     zipf = zipfile.ZipFile('UniverseFiles.zip','w', zipfile.ZIP_DEFLATED)
#     files = Filenames['path']
#     for i in range(0, len(files)):
#         final_directory = os.path.join(RootPath+"//Universe Output//", Filenames['path'][i])
#         zipf.write(final_directory, os.path.basename(final_directory))

#     zipf.close()

#     return send_file('UniverseFiles.zip',
#                      mimetype='zip',
#                      attachment_filename='Files.zip',
#                      as_attachment=True)

#Universe Extraction functions end 



#Report Extraction functions start

@app.route('/login', methods=['POST'])
def GetLoginToken():
    url = 'http://cvyhj1a18:6405/biprws/logon/long'
    login_link = 'http://www.sap.com/rws/bip'
    login_cred = {
        'username': request.json['username'],
        'password': request.json['password'],
        'auth': request.json['auth']
    }
    payload = GetPayload(login_link, login_cred['username'], login_cred['password'],login_cred['auth'])
    print(payload)
    headers = {
        'content-type': 'application/xml',
        'accept': 'application/xml'
    }
    try:
        response = requests.request("POST", url, headers=headers, data=payload)
        if response.status_code == 200:
            msg = {
                'status': "Success"
            }
        else:
            # print("Failed to generate token")
            msg = {
                'status': "Invalid username/password"
            }
    except:
        return ('Connection Error')

    return jsonify(msg)


@app.route('/extract', methods=['POST'])
def parallel():
    info = {
        'username': request.json['username'],
        'password': request.json['password'],
        'hostURL': request.json['hostURL'],
        'RepID': request.json['RepID'],
	'auth': request.json['auth']
    }

    result = Parallel(n_jobs=num_cores - 1)(delayed(drect)(info['username'], info['password'], info['hostURL'] 
                                                           , info['RepID'][i],info['auth']) for i in range(0, len(info['RepID'])))

    return jsonify({'Response': result, 'status': 'SUCCESS'})

    end_time = datetime.now()
    print("Start time - %s" % (start_time))
    print("End time - %s" % (end_time))


@app.route('/getReportOptions', methods=['GET'])
def getReportOptions():
    excelDF = pd.read_excel('./Source_Report_List.xlsx')

    reportNames = excelDF["ReportName"];
    reportCUIDs = excelDF["ReportCUID"];
    reportIDs = excelDF["ReportID"];

    result = [];

    for i in excelDF.index:
        result.append({'ReportName': reportNames[i], 'ReportCUID': reportCUIDs[i], 'ReportID': str(reportIDs[i])})

    return jsonify({'reportOptions': result})



@app.route('/download/<path:filename>', methods=['GET', 'POST'])
def download(filename):
    return send_from_directory(directory=os.getcwd(), filename=filename,
                               as_attachment=True)


@app.route('/downloadAll', methods=['POST'])
def download_all():
    location = {
        'path': request.json['path']
    }
    zipf = zipfile.ZipFile('Files.zip', 'w', zipfile.ZIP_DEFLATED)
    files = location['path']
    for i in range(0, len(files)):
        final_directory = os.path.join(os.getcwd(), location['path'][i].replace("/", "\\"))
        print("files", final_directory)
        zipf.write(final_directory, os.path.basename(final_directory))

    zipf.close()

    return send_file('Files.zip',
                     mimetype='zip',
                     attachment_filename='Files.zip',
                     as_attachment=True)


@app.route('/CheckSQL', methods=['GET'])
def getDropdown():
    folder_list = []
    if os.path.exists('output/'):
        current_directory = os.getcwd()
        final_directory = os.path.join(current_directory, 'output\\')
        files = glob2.glob(final_directory + '/**/*_VALIDATED.txt')

        if files:
            for file in files:
                print(file)
                path = os.path.dirname(file)
                folder = {
                    'folder_name': os.path.basename(path)
                }
                folder_list.append(folder)
                doc_list = [dict(t) for t in {tuple(d.items()) for d in folder_list}]

    return jsonify({'SQL_Validated_Files': doc_list})


@app.route('/CheckTWBX', methods=['GET'])
def getDropdowntwbx():
    folder_list_twbx = []
    if os.path.exists('output/'):
        current_directory = os.getcwd()
        final_directory = os.path.join(current_directory, 'output\\')
        files = glob2.glob(final_directory + '/**/*.twbx')

        if files:
            for file in files:
                print(file)
                path = os.path.dirname(file)
                folder = {
                    'folder_name': os.path.basename(path)
                }
                folder_list_twbx.append(folder)

    return jsonify({'TWBX_Files': folder_list_twbx})


@app.route('/runBot', methods=['GET', 'POST'])
def bot_executor():

    location = {
        'file': request.json['file']
    }

    temp_path = "Report_Status_Template.xlsx"
    temp_wb = openpyxl.load_workbook(temp_path)

    temp_sheet = temp_wb["UI_User_Input"]
    row_counter = 2

    files = location['file']
    for folder in files:
        temp_sheet.cell(row=row_counter, column=1).value = folder
        row_counter = row_counter + 1

    temp_wb.save('User_Input_from_UI.xlsx')
    temp_wb.close()

    PROCESS_NAME = r"C:\Users\abhisher\Documents\DE-RECT\BOTab-MultiDP_24thMarch\Start_Tableau.xaml"
    EXECUTOR_PATH = r"C:\Users\abhisher\AppData\Local\UiPath\app-20.10.6\UiRobot.exe"

    process_names = (process_names.name() for process_names in psutil.process_iter())
    if "UiPath.Executor.exe" not in process_names:
        try:
            os.system('{0} execute -f {1}'.format(EXECUTOR_PATH, PROCESS_NAME))
        except:
            msg = "Unable to start BOT"
            pass
    else:
        msg = "BOT running already"

    if os.path.exists('Report_Generation_Status.xlsx'):
        msg = "BOT ran successfully"

        wb = xlrd.open_workbook('Report_Generation_Status.xlsx')
        sh = wb.sheet_by_index(0)

        data_list = []
        for rownum in range(1, sh.nrows):
            data = OrderedDict()
            row_values = sh.row_values(rownum)
            data['Report_Name'] = row_values[0]
            data['Status'] = row_values[1]
            data_list.append(data)

        final_directory = os.path.join(os.getcwd(), 'output\\')
        file_path = glob2.glob(final_directory + '/**/*.twbx')
        folder_twbx = []

        if file_path:
            for loc_file in file_path:
                file = os.path.relpath(loc_file, os.getcwd())
                print(file)
                path = os.path.dirname(file)
                folder = {
                    os.path.basename(path): file.replace("\\", "/")
                }
                for data in data_list:
                    if data["Report_Name"] == os.path.basename(path):
                        data["file_name"] = file.replace("\\", "/")
                folder_twbx.append(folder)

    else:
        msg = "BOT failed"
        return ({'Message': msg})

    # if os.path.exists('Report_Generation_Status.xlsx'):
    #     os.remove('Report_Generation_Status.xlsx')

    return jsonify({'Message': msg, 'Response': data_list, 'TWBX_FilePath': folder_twbx})


if __name__ == '__main__':
    start_time = datetime.now()
    num_cores = multiprocessing.cpu_count()

    app.run(host='0.0.0.0', port=5000, debug=True)