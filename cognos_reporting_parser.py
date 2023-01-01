import xml.etree.ElementTree as ET
import openpyxl
import re
import os

inp = input("Enter the file name : ")
print(inp)
final_directory = os.path.join(os.getcwd(), 'Updated\\')
files = final_directory + inp
tree = ET.parse(files)
root = tree.getroot()
tag = root.tag
att = root.attrib
print(root,'\n',tag,'\n',att)

temp_path = "Template.xlsx"
temp_wb = openpyxl.load_workbook(temp_path)

for child in root:
    #Report Detail Tab
    if child.tag == '{http://developer.cognos.com/schemas/report/16.1/}reportName':
        # print("Text",child.text)
        child_text = child.text

        temp_sheet = temp_wb["Report Detail"]
        row_counter = 2

        temp_sheet.cell(row=row_counter, column=1).value = child_text
        row_counter = row_counter + 1

    if child.tag == '{http://developer.cognos.com/schemas/report/16.1/}modelPath':
        child_text = child.text

        temp_sheet = temp_wb["Report Detail"]
        row_counter = 2

        #pattern = "/package[@name='(.+?)']"
        #substring = re.search(pattern, child_text).group(1)


        if "/package[@name='" in child_text:
            print("START",child_text.find("/package[@name='"))
            start = child_text.find("/package[@name='") + len("/package[@name='")
            end = child_text.find("']/model")
            print("end",end)
            substring = child_text[start:end]
            print("substring",substring)

        temp_sheet.cell(row=row_counter, column=2).value = child.get('type')
        if "/package[@name='" in child_text:
            temp_sheet.cell(row=row_counter, column=3).value = substring

        temp_sheet.cell(row=row_counter, column=4).value = child_text.partition('/model')[0]
        row_counter = row_counter + 1

    #Object Tab
    if child.tag == '{http://developer.cognos.com/schemas/report/16.1/}layouts':
        temp_sheet = temp_wb["Object"]
        row_counter = 2

        for elem in child.iter('{http://developer.cognos.com/schemas/report/16.1/}list'):

            temp_sheet.cell(row=row_counter, column=1).value = elem.tag.partition('list')[1]
            temp_sheet.cell(row=row_counter, column=2).value = elem.attrib["name"]
            temp_sheet.cell(row=row_counter, column=3).value = elem.attrib["refQuery"]

            for elemchild in elem.iter('{http://developer.cognos.com/schemas/report/16.1/}listColumns'):
                for elemchild_1 in elemchild.iter('{http://developer.cognos.com/schemas/report/16.1/}dataItemValue'):
                    temp_sheet.cell(row=row_counter, column=4).value = elemchild_1.attrib["refDataItem"]
                    sort_yes_temp = elemchild_1.attrib["refDataItem"]
                    for elemchild_2 in elem.iter('{http://developer.cognos.com/schemas/report/16.1/}sortList'):
                        for elemchild_3 in elemchild_2.iter('{http://developer.cognos.com/schemas/report/16.1/}sortItem'):
                            if elemchild_3.attrib["refDataItem"] == sort_yes_temp:
                                temp_sheet.cell(row=row_counter, column=5).value = "Yes"
                    for elemchild_4 in elem.iter('{http://developer.cognos.com/schemas/report/16.1/}GroupbyList'):
                        for elemchild_5 in elemchild_4.iter('{http://developer.cognos.com/schemas/report/16.1/}GroupbyItem'):
                            if elemchild_5.attrib["refDataItem"] == sort_yes_temp:
                                temp_sheet.cell(row=row_counter, column=6).value = "Yes"
                    row_counter = row_counter + 1


        for elem in child.iter('{http://developer.cognos.com/schemas/report/16.1/}crosstab'):

            temp_sheet.cell(row=row_counter, column=1).value = elem.tag.partition('crosstab')[1]
            temp_sheet.cell(row=row_counter, column=2).value = elem.attrib["name"]
            temp_sheet.cell(row=row_counter, column=3).value = elem.attrib["refQuery"]

            for elemchild_1 in elem.iter('{http://developer.cognos.com/schemas/report/16.1/}crosstabNodeMember'):
                temp_sheet.cell(row=row_counter, column=4).value = elemchild_1.attrib["refDataItem"]
                sort_yes_temp = elemchild_1.attrib["refDataItem"]
                for elemchild_3 in elemchild_1.iter('{http://developer.cognos.com/schemas/report/16.1/}sortItem'):
                    if elemchild_3.attrib["refDataItem"] == sort_yes_temp:
                        temp_sheet.cell(row=row_counter, column=5).value = "Yes"
                for elemchild_4 in elemchild_1.iter('{http://developer.cognos.com/schemas/report/16.1/}GroupByItem'):
                    if elemchild_4.attrib["refDataItem"] == sort_yes_temp:
                        temp_sheet.cell(row=row_counter, column=6).value = "Yes"
                row_counter = row_counter + 1


        for elem in child.iter('{http://developer.cognos.com/schemas/report/16.1/}vizControl'):

            if "name" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=1).value = elem.attrib["name"]
            if "name" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=2).value = elem.attrib["name"]
            if "refQuery" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=3).value = elem.attrib["refQuery"]

            for elemchild_1 in elem.iter('{http://developer.cognos.com/schemas/report/16.1/}vcSlotDsColumn'):
                temp_sheet.cell(row=row_counter, column=4).value = elemchild_1.attrib["refDsColumn"]
                sort_yes_temp = elemchild_1.attrib["refDsColumn"]
                for elemchild_3 in elemchild_1.iter('{http://developer.cognos.com/schemas/report/16.1/}sortItem'):
                    if elemchild_3.attrib["refDsColumn"] == sort_yes_temp:
                        temp_sheet.cell(row=row_counter, column=5).value = "Yes"
                row_counter = row_counter + 1

        #row_counter = row_counter + 1

    #Queries Tab
    if child.tag == '{http://developer.cognos.com/schemas/report/16.1/}queries':
        temp_sheet = temp_wb["Queries"]
        row_counter = 2
        for elem in child.iter('{http://developer.cognos.com/schemas/report/16.1/}query'):
            temp_sheet.cell(row=row_counter, column=1).value = elem.attrib["name"]
            for elem_2 in elem.iter('{http://developer.cognos.com/schemas/report/16.1/}dataItem'):
                temp_sheet.cell(row=row_counter, column=2).value = elem_2.attrib["name"]
                if elem.iter('{http://developer.cognos.com/schemas/report/16.1/}joinFilter'):
                    for elem_4 in elem.iter('{http://developer.cognos.com/schemas/report/16.1/}source'):
                        for elem_5 in elem_4.iter('{http://developer.cognos.com/schemas/report/16.1/}joinOperation'):
                            for elem_6 in elem_5.iter('{http://developer.cognos.com/schemas/report/16.1/}joinFilter'):
                                for elem_7 in elem_6.iter(
                                        '{http://developer.cognos.com/schemas/report/16.1/}filterExpression'):
                                    temp_sheet.cell(row=row_counter, column=10).value = elem_7.text
                if elem.iter('{http://developer.cognos.com/schemas/report/16.1/}joinOperand'):
                    for elem_8 in elem.iter('{http://developer.cognos.com/schemas/report/16.1/}joinOperand'):
                        for elem_9 in elem_8.iter(
                                '{http://developer.cognos.com/schemas/report/16.1/}queryRef'):
                            if temp_sheet.cell(row=row_counter, column=11).value == None:
                                temp_sheet.cell(row=row_counter, column=11).value = ""
                            if "cardinality" in elem_8.attrib.keys() and "refQuery" in elem_9.attrib.keys():
                                #temp_value = str(temp_sheet.cell(row=row_counter, column=11))
                                temp_sheet.cell(row=row_counter, column=11).value += "Cardinality =" + \
                                          elem_8.attrib[
                                              "cardinality"] + "-" + \
                                          elem_9.attrib["refQuery"] + "\n"

                #For Union Queries in Queries Tab
                if elem.iter('{http://developer.cognos.com/schemas/report/16.1/}queryOperation'):
                    temp_value = ""
                    for elem_10 in elem.iter('{http://developer.cognos.com/schemas/report/16.1/}queryOperation'):
                        if "Union" in elem_10.attrib["name"]:
                            for elem_11 in elem.iter('{http://developer.cognos.com/schemas/report/16.1/}queryRefs'):
                                for elem_12 in elem.iter(
                                        '{http://developer.cognos.com/schemas/report/16.1/}queryRef'):
                                    temp_value += elem_12.attrib["refQuery"] + ','
                            temp_sheet.cell(row=row_counter, column=9).value = temp_value[0:-1]

                #For Intersection Queries in Queries Tab
                if elem.iter('{http://developer.cognos.com/schemas/report/16.1/}queryOperation'):
                    temp_value = ""
                    for elem_10 in elem.iter('{http://developer.cognos.com/schemas/report/16.1/}queryOperation'):
                        if "Intersection" in elem_10.attrib["name"]:
                            for elem_11 in elem.iter('{http://developer.cognos.com/schemas/report/16.1/}queryRefs'):
                                for elem_12 in elem.iter(
                                        '{http://developer.cognos.com/schemas/report/16.1/}queryRef'):
                                    temp_value += elem_12.attrib["refQuery"] + ','
                            temp_sheet.cell(row=row_counter, column=9).value = temp_value[0:-1]

                row_counter = row_counter + 1



        row_counter = 2
        for elem in child.iter('{http://developer.cognos.com/schemas/report/16.1/}query'):
            for elem_1 in elem.iter('{http://developer.cognos.com/schemas/report/16.1/}expression'):
                temp_sheet.cell(row=row_counter, column=4).value = elem_1.text
                row_counter = row_counter + 1


        row_counter = 2
        for elem in child.iter('{http://developer.cognos.com/schemas/report/16.1/}dataItem'):
            if "aggregate" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=5).value = elem.attrib["aggregate"]
            if "rollupAggregate" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=6).value = elem.attrib["rollupAggregate"]
            if "sort" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=7).value = elem.attrib["sort"]
            if "GroupBy" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=8).value = elem.attrib["GroupBy"]
            row_counter = row_counter + 1


    #Prompts Tab
    if child.tag == '{http://developer.cognos.com/schemas/report/16.1/}layouts':
        temp_sheet = temp_wb["Prompts"]
        row_counter = 2
        for elem in child.iter('{http://developer.cognos.com/schemas/report/16.1/}selectValue'):
            if "name" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=1).value = elem.attrib["name"]
            temp_sheet.cell(row=row_counter, column=2).value = "Value Prompt"
            if "refQuery" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=3).value = elem.attrib["refQuery"]
            if "parameter" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=4).value = elem.attrib["parameter"]
            if "multiSelect" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=5).value = elem.attrib["multiSelect"]
            if "required" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=6).value = elem.attrib["required"]
            if "cascadeOn" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=7).value = elem.attrib["cascadeOn"]
            row_counter = row_counter + 1

        for elem in child.iter('{http://developer.cognos.com/schemas/report/16.1/}selectWithSearch'):
            if "name" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=1).value = elem.attrib["name"]
            temp_sheet.cell(row=row_counter, column=2).value = "Select and Search Prompt"
            if "refQuery" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=3).value = elem.attrib["refQuery"]
            if "parameter" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=4).value = elem.attrib["parameter"]
            if "multiSelect" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=5).value = elem.attrib["multiSelect"]
            if "required" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=6).value = elem.attrib["required"]
            if "cascadeOn" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=7).value = elem.attrib["cascadeOn"]
            row_counter = row_counter + 1

        for elem in child.iter('{http://developer.cognos.com/schemas/report/16.1/}selectDate'):
            if "name" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=1).value = elem.attrib["name"]
            temp_sheet.cell(row=row_counter, column=2).value = "Date Prompt"
            if "refQuery" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=3).value = elem.attrib["refQuery"]
            if "parameter" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=4).value = elem.attrib["parameter"]
            if "multiSelect" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=5).value = elem.attrib["multiSelect"]
            if "required" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=6).value = elem.attrib["required"]
            if "cascadeOn" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=7).value = elem.attrib["cascadeOn"]
            row_counter = row_counter + 1

        for elem in child.iter('{http://developer.cognos.com/schemas/report/16.1/}selectDateTime'):
            if "name" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=1).value = elem.attrib["name"]
            temp_sheet.cell(row=row_counter, column=2).value = "Date Time Prompt"
            if "refQuery" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=3).value = elem.attrib["refQuery"]
            if "parameter" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=4).value = elem.attrib["parameter"]
            if "multiSelect" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=5).value = elem.attrib["multiSelect"]
            if "required" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=6).value = elem.attrib["required"]
            if "cascadeOn" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=7).value = elem.attrib["cascadeOn"]
            row_counter = row_counter + 1

        for elem in child.iter('{http://developer.cognos.com/schemas/report/16.1/}textBox'):
            if "name" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=1).value = elem.attrib["name"]
            temp_sheet.cell(row=row_counter, column=2).value = "Textbox Prompt"
            if "refQuery" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=3).value = elem.attrib["refQuery"]
            if "parameter" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=4).value = elem.attrib["parameter"]
            if "multiSelect" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=5).value = elem.attrib["multiSelect"]
            if "required" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=6).value = elem.attrib["required"]
            if "cascadeOn" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=7).value = elem.attrib["cascadeOn"]
            row_counter = row_counter + 1

        for elem in child.iter('{http://developer.cognos.com/schemas/report/16.1/}selectTime'):
            if "name" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=1).value = elem.attrib["name"]
            temp_sheet.cell(row=row_counter, column=2).value = "Time Prompt"
            if "refQuery" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=3).value = elem.attrib["refQuery"]
            if "parameter" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=4).value = elem.attrib["parameter"]
            if "multiSelect" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=5).value = elem.attrib["multiSelect"]
            if "required" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=6).value = elem.attrib["required"]
            if "cascadeOn" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=7).value = elem.attrib["cascadeOn"]
            row_counter = row_counter + 1

    #Filter Tab
    if child.tag == '{http://developer.cognos.com/schemas/report/16.1/}queries':
        temp_sheet = temp_wb["Filter"]
        row_counter = 2
        for elem in child.iter('{http://developer.cognos.com/schemas/report/16.1/}query'):
            temp_sheet.cell(row=row_counter, column=1).value = elem.attrib["name"]
            if elem.iter('{http://developer.cognos.com/schemas/report/16.1/}filterExpression'):
                for elem_2 in elem.iter('{http://developer.cognos.com/schemas/report/16.1/}filterExpression'):
                    temp_sheet.cell(row=row_counter, column=2).value = elem_2.text
                    temp = elem_2.text
                    if len(temp) > 0 and elem.iter('{http://developer.cognos.com/schemas/report/16.1/}SummaryFilters'):
                        temp_sheet.cell(row=row_counter, column=3).value = "Summary Filters"

                    if len(temp) > 0 and elem.iter('{http://developer.cognos.com/schemas/report/16.1/}DetailFilters'):
                        temp_sheet.cell(row=row_counter, column=3).value = "Detail Filters"

                    row_counter = row_counter + 1


        row_counter = 2
        for elem in child.iter('{http://developer.cognos.com/schemas/report/16.1/}queries'):
            for elem_1 in elem.iter('{http://developer.cognos.com/schemas/report/16.1/}summaryFilter'):
                if "use" in elem_1.attrib.keys():
                    temp_sheet.cell(row=row_counter, column=4).value = elem_1.attrib["use"]
                    row_counter = row_counter + 1

            for elem_2 in elem.iter('{http://developer.cognos.com/schemas/report/16.1/}detailFilter'):
                if "use" in elem_2.attrib.keys():
                    temp_sheet.cell(row=row_counter, column=4).value = elem_2.attrib["use"]
                    row_counter = row_counter + 1

    #Variables Tab
    if child.tag == '{http://developer.cognos.com/schemas/report/16.1/}reportVariables':
        temp_sheet = temp_wb["Variables"]
        row_counter = 2
        for elem in child.iter('{http://developer.cognos.com/schemas/report/16.1/}reportVariable'):
            if "name" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=1).value = elem.attrib["name"]
            if "type" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=2).value = elem.attrib["type"]
            row_counter = row_counter + 1

        row_counter = 2
        for elem_2 in child.iter('{http://developer.cognos.com/schemas/report/16.1/}reportExpression'):
            temp_sheet.cell(row=row_counter, column=3).value = elem_2.text
            row_counter = row_counter + 1

    #Drill Through Tab
    if child.tag == '{http://developer.cognos.com/schemas/report/16.1/}layouts':
        temp_sheet = temp_wb["Drill Through"]
        row_counter = 2
        for elem in child.iter('{http://developer.cognos.com/schemas/report/16.1/}reportDrill'):
            if "name" in elem.attrib.keys():
                temp_sheet.cell(row=row_counter, column=1).value = elem.attrib["name"]

            for elem_1 in elem.iter('{http://developer.cognos.com/schemas/report/16.1/}reportPath'):
                if "path" in elem_1.attrib.keys():
                    temp_sheet.cell(row=row_counter, column=2).value = elem_1.attrib["path"]

            for elem_2 in elem.iter('{http://developer.cognos.com/schemas/report/16.1/}XMLAttribute'):
                if elem_2.attrib["name"] == "ReportName":
                    temp_sheet.cell(row=row_counter, column=3).value = elem_2.attrib["value"]

            for elem_3 in elem.iter('{http://developer.cognos.com/schemas/report/16.1/}drillLinks'):
                for elem_4 in elem_3.iter('{http://developer.cognos.com/schemas/report/16.1/}parameterContext'):
                    if "parameter" in elem_4.attrib.keys():
                        temp_sheet.cell(row=row_counter, column=4).value = elem_4.attrib["parameter"]
                        row_counter = row_counter + 1


temp_wb.save('Output.xlsx')
temp_wb.close()
